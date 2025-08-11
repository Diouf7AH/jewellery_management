from io import BytesIO

import qrcode
from django.db import transaction
from django.http import HttpResponse
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
# from knox.auth import TokenAuthentication
from rest_framework import status
from rest_framework.parsers import (FileUploadParser, FormParser,
                                    MultiPartParser)
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from rest_framework.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.http import Http404 

from rest_framework.parsers import JSONParser

from backend.renderers import UserRenderer
from stock.serializers import StockSerializer
from store.models import Bijouterie, Categorie, Marque, Modele, Produit, Gallery, Purete, CategorieModele, MarquePurete
from store.serializers import (BijouterieSerializer, CategorieSerializer,
                            ModeleSerializer, ModeleListSerializer,
                            ProduitSerializer, PureteSerializer, GallerySerializer, ProduitWithGallerySerializer, CategorieWithModelesSerializer, MarqueCreateSerializer)

#Export to excel
import qrcode
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

allowed_roles_admin_manager_vendeur = ['admin', 'manager', 'vendeur']
allowed_roles_admin_manager = ['admin', 'manager']

# Create your views here.
class BijouterieListAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Lister les bijouteries",
        operation_description="Retourne la liste de toutes les bijouteries enregistrées. Accès réservé aux rôles : admin, manager.",
        responses={
            200: openapi.Response(
                description="Liste des bijouteries",
                schema=BijouterieSerializer(many=True)
            ),
            403: "Accès refusé"
        }
    )
    def get(self, request):

        user_role = getattr(request.user.user_role, 'role', None)
        if user_role not in allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        bijouteries = Bijouterie.objects.all()
        serializer = BijouterieSerializer(bijouteries, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class BijouterieCreateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(
        operation_summary="Créer une bijouterie",
        operation_description="Permet à un administrateur ou manager d’enregistrer une nouvelle bijouterie, y compris les logos et les contacts.",
        request_body=BijouterieSerializer,
        responses={
            201: openapi.Response("Bijouterie créée avec succès", BijouterieSerializer),
            400: openapi.Response("Requête invalide")
        }
    )
    def post(self, request):
        user_role = getattr(request.user.user_role, 'role', None)

        if user_role not in allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        serializer = BijouterieSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class BijouterieUpdateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @swagger_auto_schema(
        operation_summary="Mettre à jour une bijouterie",
        operation_description="Permet de modifier les informations d'une bijouterie existante.",
        manual_parameters=[
            openapi.Parameter(
                'pk', openapi.IN_PATH, description="ID de la bijouterie à mettre à jour",
                type=openapi.TYPE_INTEGER, required=True
            )
        ],
        request_body=BijouterieSerializer,
        responses={
            200: openapi.Response("Bijouterie mise à jour", BijouterieSerializer),
            400: "Requête invalide",
            403: "Accès refusé",
            404: "Bijouterie introuvable"
        }
    )
    def put(self, request, pk):
        if not request.user.user_role or request.user.user_role.role not in allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        try:
            bijouterie = Bijouterie.objects.get(pk=pk)
        except Bijouterie.DoesNotExist:
            return Response({"detail": "Bijouterie introuvable."}, status=404)

        serializer = BijouterieSerializer(bijouterie, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=200)
        return Response(serializer.errors, status=400)

    def patch(self, request, pk):
        if not request.user.user_role or request.user.user_role.role not in allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        try:
            bijouterie = Bijouterie.objects.get(pk=pk)
        except Bijouterie.DoesNotExist:
            return Response({"detail": "Bijouterie introuvable."}, status=404)

        serializer = BijouterieSerializer(bijouterie, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=200)
        return Response(serializer.errors, status=400)


class BijouterieDeleteAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Supprimer une bijouterie",
        operation_description="Supprime définitivement une bijouterie à partir de son ID.",
        manual_parameters=[
            openapi.Parameter(
                'pk', openapi.IN_PATH, description="ID de la bijouterie à supprimer",
                type=openapi.TYPE_INTEGER, required=True
            )
        ],
        responses={
            204: openapi.Response("Bijouterie supprimée avec succès"),
            403: "Accès refusé",
            404: "Bijouterie introuvable"
        }
    )
    def delete(self, request, pk):
        if not request.user.user_role or request.user.user_role.role not in allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        try:
            bijouterie = Bijouterie.objects.get(pk=pk)
            bijouterie.delete()
            return Response({"message": "Bijouterie supprimée avec succès."}, status=status.HTTP_204_NO_CONTENT)
        except Bijouterie.DoesNotExist:
            return Response({"detail": "Bijouterie introuvable."}, status=status.HTTP_404_NOT_FOUND)


class CategorieListAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser]

    @swagger_auto_schema(
        operation_description="Lister les catégories avec filtrage par nom (paramètre `search`).",
        manual_parameters=[
            openapi.Parameter(
                'search', openapi.IN_QUERY,
                description="Filtrer par nom de catégorie",
                type=openapi.TYPE_STRING
            )
        ],
        responses={200: openapi.Response('Liste des catégories', CategorieSerializer(many=True))}
    )
    def get(self, request):
        role = getattr(request.user.user_role, 'role', None)
        if role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=403)

        search_query = request.query_params.get('search')
        categories = Categorie.objects.all()

        if search_query:
            categories = categories.filter(nom__icontains=search_query)

        serializer = CategorieSerializer(categories, many=True)
        return Response(serializer.data)


class CategorieCreateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    parser_classes = (FormParser, MultiPartParser, FileUploadParser)

    @swagger_auto_schema(
        operation_description="Créer une nouvelle catégorie avec un nom et une image.",
        request_body=CategorieSerializer,
        responses={
            status.HTTP_201_CREATED: openapi.Response('Catégorie créée avec succès', CategorieSerializer),
            status.HTTP_400_BAD_REQUEST: openapi.Response('Requête invalide')
        }
    )
    def post(self, request):
        role = getattr(request.user.user_role, 'role', None)
        if role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        serializer = CategorieSerializer(data=request.data)
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

# class CategorieUpdateAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]
#     parser_classes = (FormParser, MultiPartParser, FileUploadParser)

#     def get_object(self, pk):
#         try:
#             return Categorie.objects.get(pk=pk)
#         except Categorie.DoesNotExist:
#             return None

#     def has_access(self, user):
#         return user.user_role and user.user_role.role in ['admin', 'manager']

#     @swagger_auto_schema(
#         operation_summary="Modifier une catégorie (PUT)",
#         operation_description="Remplace complètement une catégorie existante.",
#         request_body=CategorieSerializer,
#         responses={
#             200: openapi.Response("Catégorie mise à jour avec succès", CategorieSerializer),
#             400: "Erreur de validation",
#             403: "Accès refusé",
#             404: "Catégorie non trouvée"
#         }
#     )
#     def put(self, request, pk):
#         if not self.has_access(request.user):
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         categorie = self.get_object(pk)
#         if not categorie:
#             return Response({"message": "Catégorie non trouvée"}, status=status.HTTP_404_NOT_FOUND)

#         serializer = CategorieSerializer(categorie, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#     @swagger_auto_schema(
#         operation_summary="Modifier partiellement une catégorie (PATCH)",
#         operation_description="Met à jour partiellement les champs d'une catégorie existante.",
#         request_body=CategorieSerializer,
#         responses={
#             200: openapi.Response("Catégorie mise à jour avec succès", CategorieSerializer),
#             400: "Erreur de validation",
#             403: "Accès refusé",
#             404: "Catégorie non trouvée"
#         }
#     )
#     def patch(self, request, pk):
#         if not self.has_access(request.user):
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         categorie = self.get_object(pk)
#         if not categorie:
#             return Response({"message": "Catégorie non trouvée"}, status=status.HTTP_404_NOT_FOUND)

#         serializer = CategorieSerializer(categorie, data=request.data, partial=True)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


class CategorieUpdateByNameAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    parser_classes = (FormParser, MultiPartParser, FileUploadParser)

    def get_object(self, nom):
        try:
            return Categorie.objects.get(nom__iexact=nom)
        except Categorie.DoesNotExist:
            return None

    def has_access(self, user):
        return user.user_role and user.user_role.role in ['admin', 'manager']

    @swagger_auto_schema(
        operation_summary="Remplacer une catégorie par son nom (PUT)",
        operation_description="Remplace entièrement une catégorie identifiée par son nom.",
        manual_parameters=[
            openapi.Parameter(
                'nom', openapi.IN_PATH,
                description="Nom exact de la catégorie (insensible à la casse)",
                type=openapi.TYPE_STRING
            )
        ],
        request_body=CategorieSerializer,
        responses={
            200: openapi.Response("Catégorie mise à jour avec succès", CategorieSerializer),
            400: "Erreur de validation",
            403: "Accès refusé",
            404: "Catégorie non trouvée"
        }
    )
    def put(self, request, nom):
        if not self.has_access(request.user):
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        categorie = self.get_object(nom)
        if not categorie:
            return Response({"message": "Catégorie non trouvée"}, status=status.HTTP_404_NOT_FOUND)

        serializer = CategorieSerializer(categorie, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(
        operation_summary="Mettre à jour partiellement une catégorie par son nom (PATCH)",
        operation_description="Met à jour partiellement une catégorie identifiée par son nom.",
        manual_parameters=[
            openapi.Parameter(
                'nom', openapi.IN_PATH,
                description="Nom exact de la catégorie (insensible à la casse)",
                type=openapi.TYPE_STRING
            )
        ],
        request_body=CategorieSerializer,
        responses={
            200: openapi.Response("Catégorie mise à jour partiellement", CategorieSerializer),
            400: "Erreur de validation",
            403: "Accès refusé",
            404: "Catégorie non trouvée"
        }
    )
    def patch(self, request, nom):
        if not self.has_access(request.user):
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        categorie = self.get_object(nom)
        if not categorie:
            return Response({"message": "Catégorie non trouvée"}, status=status.HTTP_404_NOT_FOUND)

        serializer = CategorieSerializer(categorie, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class CategorieDeleteAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Supprimer une catégorie",
        operation_description="Supprime une catégorie existante par son ID.",
        manual_parameters=[
            openapi.Parameter(
                'pk',
                openapi.IN_PATH,
                description="ID de la catégorie à supprimer",
                type=openapi.TYPE_INTEGER,
                required=True
            )
        ],
        responses={
            204: openapi.Response(description="Catégorie supprimée avec succès"),
            403: "Accès refusé",
            404: "Catégorie non trouvée"
        }
    )
    def delete(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        try:
            categorie = Categorie.objects.get(pk=pk)
        except Categorie.DoesNotExist:
            return Response({"message": "Catégorie non trouvée"}, status=status.HTTP_404_NOT_FOUND)

        categorie.delete()
        return Response({"message": "Catégorie supprimée avec succès."}, status=status.HTTP_204_NO_CONTENT)


class PureteListAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Lister les puretés",
        operation_description="Retourne la liste de toutes les puretés, avec option de filtrage par valeur partielle (`?purete=`).",
        manual_parameters=[
            openapi.Parameter(
                'purete', openapi.IN_QUERY,
                description="Recherche partielle par valeur de pureté (ex: 18 ou 21K)",
                type=openapi.TYPE_STRING
            )
        ],
        responses={200: openapi.Response('Liste des puretés', PureteSerializer(many=True))}
    )
    def get(self, request):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        search = request.query_params.get('purete')
        puretes = Purete.objects.all()

        if search:
            puretes = puretes.filter(purete__icontains=search)

        serializer = PureteSerializer(puretes, many=True)
        return Response(serializer.data)



class PureteCreateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Créer une nouvelle pureté",
        operation_description="Permet à un administrateur ou un manager d'ajouter une nouvelle pureté (ex : 18K, 24K).",
        request_body=PureteSerializer,
        responses={
            status.HTTP_201_CREATED: openapi.Response('Pureté créée avec succès', PureteSerializer),
            status.HTTP_400_BAD_REQUEST: openapi.Response('Erreur de validation')
        }
    )
    def post(self, request):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        serializer = PureteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


class PureteUpdateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Modifier une pureté",
        operation_description="Permet à un admin ou un manager de mettre à jour une pureté existante (PUT).",
        request_body=PureteSerializer,
        responses={
            200: openapi.Response("Mise à jour réussie", PureteSerializer),
            400: "Erreur de validation",
            403: "Accès refusé",
            404: "Pureté non trouvée"
        }
    )
    def put(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=403)

        purete = self.get_object(pk)
        if not purete:
            return Response({"message": "Pureté introuvable"}, status=404)

        serializer = PureteSerializer(purete, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    @swagger_auto_schema(
        operation_summary="Modifier partiellement une pureté",
        operation_description="Permet de modifier partiellement une pureté existante (PATCH).",
        request_body=PureteSerializer,
        responses={
            200: openapi.Response("Mise à jour partielle réussie", PureteSerializer),
            400: "Erreur de validation",
            403: "Accès refusé",
            404: "Pureté non trouvée"
        }
    )
    def patch(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=403)

        purete = self.get_object(pk)
        if not purete:
            return Response({"message": "Pureté introuvable"}, status=404)

        serializer = PureteSerializer(purete, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def get_object(self, pk):
        try:
            return Purete.objects.get(pk=pk)
        except Purete.DoesNotExist:
            return None


class PureteDeleteAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Supprimer une pureté",
        operation_description="Supprime une pureté par son ID. Accès réservé aux rôles admin ou manager.",
        manual_parameters=[
            openapi.Parameter(
                name='pk',
                in_=openapi.IN_PATH,
                description="ID de la pureté à supprimer",
                type=openapi.TYPE_INTEGER
            )
        ],
        responses={
            204: "Supprimée avec succès",
            403: "Accès refusé",
            404: "Pureté introuvable"
        }
    )
    def delete(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=403)

        purete = self.get_object(pk)
        if not purete:
            return Response({"detail": "Pureté introuvable"}, status=404)

        purete.delete()
        return Response({"message": "Pureté supprimée avec succès"}, status=204)

    def get_object(self, pk):
        try:
            return Purete.objects.get(pk=pk)
        except Purete.DoesNotExist:
            return None


# class MarqueListAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les marques",
#         operation_description="Récupère la liste de toutes les marques disponibles.",
#         manual_parameters=[
#             openapi.Parameter('search', openapi.IN_QUERY, description="Filtrer par nom de marque", type=openapi.TYPE_STRING)
#         ],
#         responses={200: openapi.Response('Liste des marques', MarqueSerializer(many=True))}
#     )
#     def get(self, request):
#         user = request.user
#         if not user.user_role or user.user_role.role not in ['admin', 'manager']:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         search_query = request.GET.get('search')
#         marques = Marque.objects.all()

#         if search_query:
#             marques = marques.filter(marque__icontains=search_query)

#         serializer = MarqueSerializer(marques, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)


# class MarqueListAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister toutes les marques avec leurs catégories associées",
#         responses={200: "Liste des marques"}
#     )
#     def get(self, request):
#         marques = Marque.objects.all().prefetch_related('marque_categories__categorie')
#         results = []

#         for marque in marques:
#             categories = [
#                 {
#                     "id": cm.categorie.id,
#                     "nom": cm.categorie.nom,
#                     "image": request.build_absolute_uri(cm.categorie.get_image_url())
#                 }
#                 for cm in marque.marque_categories.all()
#             ]

#             results.append({
#                 "id": marque.id,
#                 "marque": marque.marque,
#                 "prix": float(marque.prix),
#                 "purete": str(marque.purete.purete if marque.purete else ""),
#                 "categories": categories
#             })

#         return Response(results, status=200)


# class MarqueListAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les marques",
#         operation_description="Récupère la liste des marques avec filtres : nom, catégorie, pureté.",
#         manual_parameters=[
#             openapi.Parameter('search', openapi.IN_QUERY, description="Filtrer par nom de marque", type=openapi.TYPE_STRING),
#         ],
#         responses={200: openapi.Response('Liste des marques', MarqueSerializer(many=True))}
#     )
#     def get(self, request):
#         user = request.user
#         if not user.user_role or user.user_role.role not in ['admin', 'manager']:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         search_query = request.GET.get('search')

#         marques = Marque.objects.all()

#         if search_query:
#             marques = marques.filter(marque__icontains=search_query)

#         serializer = MarqueSerializer(marques, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)
    

# class MarqueCreateAPIView(APIView):
#     permission_classes = [IsAuthenticated]
#     parser_classes = [JSONParser]
    
#     # ✅ Rôles autorisés à créer une marque
#     allowed_roles_admin_manager = ['admin', 'manager']

#     @swagger_auto_schema(
#         operation_summary="Créer une nouvelle marque",
#         request_body=MarqueSerializer,
#         responses={
#             201: openapi.Response(description="Marque créée avec succès", schema=MarqueSerializer),
#             400: "Erreur de validation",
#         }
#     )
#     def post(self, request):
        
#         user = request.user
#         if not user.user_role or user.user_role.role not in self.allowed_roles_admin_manager:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         serializer = MarqueSerializer(data=request.data)
#         if serializer.is_valid():
#             marque = serializer.save()
#             return Response(MarqueSerializer(marque).data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# class CategorieCreateWithModelesView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Créer une catégorie avec des modèles (création des modèles si besoin)",
#         request_body=openapi.Schema(
#             type=openapi.TYPE_OBJECT,
#             required=["nom", "modeles"],
#             properties={
#                 "nom": openapi.Schema(type=openapi.TYPE_STRING, description="Nom de la catégorie"),
#                 "modeles": openapi.Schema(
#                     type=openapi.TYPE_ARRAY,
#                     items=openapi.Items(type=openapi.TYPE_STRING),
#                     description="Liste des noms de modèles à créer/lier"
#                 )
#             }
#         ),
#         responses={
#             201: openapi.Response(description="Catégorie et modèles enregistrés avec succès."),
#             400: "Requête invalide"
#         }
#     )
#     def post(self, request):
#         data = request.data
#         nom_categorie = data.get("nom", "").strip()
#         modeles_noms = data.get("modeles", [])

#         if not nom_categorie:
#             return Response({"error": "Le nom de la catégorie est requis."}, status=400)

#         if not isinstance(modeles_noms, list) or not modeles_noms:
#             return Response({"error": "La liste des modèles est requise."}, status=400)

#         # Création ou récupération de la catégorie
#         try:
#             categorie = Categorie.objects.get(nom=nom_categorie.title())
#         except Categorie.DoesNotExist:
#             return Response({"error": "Catégorie introuvable."}, status=404)

#         modeles_crees = []
#         liaisons_crees = []

#         for modele_nom in modeles_noms:
#             modele_nom_clean = modele_nom.strip().title()
#             if not modele_nom_clean:
#                 continue

#             modele, _ = Modele.objects.get_or_create(modele=modele_nom_clean)

#             # Lier à la catégorie
#             _, created = CategorieModele.objects.get_or_create(categorie=categorie, modele=modele)

#             modeles_crees.append(modele.modele)
#             if created:
#                 liaisons_crees.append(modele.modele)

#         return Response({
#             "message": "✅ Catégorie et modèles enregistrés.",
#             "categorie": categorie.nom,
#             "modeles_crees_ou_existant": modeles_crees,
#             "liaisons_nouvelles": liaisons_crees
#         }, status=201)


class MarqueCreateWithPureteView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Créer une marque et l’associer à des puretés existantes avec prix",
        request_body=MarqueCreateSerializer,
        responses={
            201: "✅ Marque et liaisons créées ou mises à jour",
            400: "Requête invalide",
            404: "Pureté non trouvée"
        }
    )
    def post(self, request):
        serializer = MarqueCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        marque_nom = serializer.validated_data["marque"].strip().title()
        puretes_data = serializer.validated_data["puretes"]

        # Créer ou récupérer la marque
        marque, _ = Marque.objects.get_or_create(marque=marque_nom)

        liaisons = []
        for entry in puretes_data:
            purete_id = entry["purete_id"]
            prix = entry["prix"]

            try:
                purete = Purete.objects.get(id=purete_id)
            except Purete.DoesNotExist:
                return Response(
                    {"error": f"❌ Pureté avec ID {purete_id} introuvable."},
                    status=404
                )

            # Créer ou mettre à jour la liaison Marque ↔ Purete
            mp_obj, created = MarquePurete.objects.get_or_create(
                marque=marque,
                purete=purete,
                defaults={"prix": prix}
            )

            if not created:
                mp_obj.prix = prix
                mp_obj.save()

            liaisons.append({
                "purete": purete.purete,
                "prix": str(mp_obj.prix)
            })

        return Response({
            "message": "✅ Marque et puretés associées avec succès.",
            "marque": marque.marque,
            "puretes_associees": liaisons
        }, status=201)


class MarquePutWithPureteView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        request_body=MarqueCreateSerializer,
        operation_summary="""🔁 Mettre à jour une marque avec une nouvelle liste de puretés (remplacement complet)
                            PUT = Remplacer toutes les liaisons (pas de mise à jour partielle).
                            Exemple: Si Strass est une marque existante,
                            Elle sera associée uniquement aux puretés 1 et 3 avec les prix fournis,
                            Toutes ses autres associations précédentes seront supprimées.""",
        responses={
            200: "✅ Marque et puretés mises à jour",
            400: "❌ Données invalides",
            404: "❌ Marque ou pureté introuvable"
        }
    )
    def put(self, request):
        serializer = MarqueCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        marque_nom = serializer.validated_data['marque'].strip().title()
        puretes_data = serializer.validated_data['puretes']

        try:
            marque = Marque.objects.get(marque__iexact=marque_nom)
        except Marque.DoesNotExist:
            return Response({"error": f"Marque '{marque_nom}' introuvable."}, status=404)

        # Supprimer les anciennes liaisons
        MarquePurete.objects.filter(marque=marque).delete()

        nouvelles_liaisons = []
        for entry in puretes_data:
            try:
                purete = Purete.objects.get(id=entry['purete_id'])
            except Purete.DoesNotExist:
                return Response(
                    {"error": f"Pureté avec ID {entry['purete_id']} introuvable."},
                    status=404
                )

            liaison = MarquePurete.objects.create(
                marque=marque,
                purete=purete,
                prix=entry['prix']
            )

            nouvelles_liaisons.append({
                "id": purete.id,
                "purete": purete.purete,
                "prix": str(liaison.prix)
            })

        return Response({
            "message": "✅ Liaisons marque–pureté mises à jour.",
            "marque": marque.marque,
            "puretes": nouvelles_liaisons
        }, status=200)


class MarquePatchWithPureteView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        request_body=MarqueCreateSerializer,
        operation_summary="🔧 Modifier ou ajouter des puretés pour une marque existante (sans supprimer les anciennes)",
        responses={
            200: "✅ Mise à jour ou ajout effectué avec succès",
            400: "❌ Données invalides",
            404: "❌ Marque ou pureté introuvable"
        }
    )
    def patch(self, request):
        serializer = MarqueCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        marque_nom = serializer.validated_data['marque'].strip().title()
        puretes_data = serializer.validated_data['puretes']

        try:
            marque = Marque.objects.get(marque__iexact=marque_nom)
        except Marque.DoesNotExist:
            return Response({"error": f"Marque '{marque_nom}' introuvable."}, status=404)

        mises_a_jour = []

        for entry in puretes_data:
            try:
                purete = Purete.objects.get(id=entry['purete_id'])
            except Purete.DoesNotExist:
                return Response(
                    {"error": f"Pureté avec ID {entry['purete_id']} introuvable."},
                    status=404
                )

            # Créer ou mettre à jour la liaison Marque–Pureté
            liaison, created = MarquePurete.objects.get_or_create(
                marque=marque,
                purete=purete,
                defaults={'prix': entry['prix']}
            )

            if not created:
                liaison.prix = entry['prix']
                liaison.save()

            mises_a_jour.append({
                "id": purete.id,
                "purete": purete.purete,
                "prix": str(liaison.prix),
                "etat": "ajouté" if created else "modifié"
            })

        return Response({
            "message": "✅ Mise à jour partielle réussie.",
            "marque": marque.marque,
            "puretés_ajoutées_ou_modifiées": mises_a_jour
        }, status=200)



class CategorieCreateWithModelesView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Associer des modèles à une catégorie existante",
        request_body=CategorieWithModelesSerializer,
        responses={
            201: openapi.Response(description="Modèles liés avec succès."),
            400: "Requête invalide",
            404: "Catégorie introuvable"
        }
    )
    def post(self, request):
        serializer = CategorieWithModelesSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        nom_categorie = serializer.validated_data["nom"].strip().title()
        modeles_noms = serializer.validated_data["modeles"]

        try:
            categorie = Categorie.objects.get(nom=nom_categorie)
        except Categorie.DoesNotExist:
            return Response({"error": "Catégorie introuvable."}, status=404)

        modeles_crees = []
        liaisons_nouvelles = []

        for nom in modeles_noms:
            modele, _ = Modele.objects.get_or_create(modele=nom)
            _, created = CategorieModele.objects.get_or_create(categorie=categorie, modele=modele)

            modeles_crees.append(modele.modele)
            if created:
                liaisons_nouvelles.append(modele.modele)

        return Response({
            "message": "✅ Catégorie et modèles enregistrés.",
            "categorie": categorie.nom,
            "modeles_crees_ou_existant": modeles_crees,
            "liaisons_nouvelles": liaisons_nouvelles
        }, status=201)


# class CategorieUpdateModelesView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Remplacer tous les modèles d'une catégorie (PUT)",
#         request_body=CategorieWithModelesSerializer,
#         responses={200: "Modèles remplacés avec succès", 400: "Erreur", 404: "Catégorie non trouvée"}
#     )
#     def put(self, request):
#         serializer = CategorieWithModelesSerializer(data=request.data)
#         if not serializer.is_valid():
#             return Response(serializer.errors, status=400)

#         nom_categorie = serializer.validated_data["nom"].strip().title()
#         modeles_noms = serializer.validated_data["modeles"]

#         try:
#             categorie = Categorie.objects.get(nom=nom_categorie)
#         except Categorie.DoesNotExist:
#             return Response({"error": "Catégorie introuvable."}, status=404)

#         # Supprimer les liaisons existantes
#         CategorieModele.objects.filter(categorie=categorie).delete()

#         modeles_ajoutes = []
#         for nom in modeles_noms:
#             modele, _ = Modele.objects.get_or_create(modele=nom)
#             CategorieModele.objects.create(categorie=categorie, modele=modele)
#             modeles_ajoutes.append(modele.modele)

#         return Response({
#             "message": "✅ Modèles remplacés avec succès.",
#             "categorie": categorie.nom,
#             "nouveaux_modeles": modeles_ajoutes
#         }, status=200)


# class CategoriePatchModelesView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Ajouter de nouveaux modèles à une catégorie (PATCH)",
#         request_body=CategorieWithModelesSerializer,
#         responses={200: "Modèles ajoutés", 400: "Erreur", 404: "Catégorie non trouvée"}
#     )
#     def patch(self, request):
#         serializer = CategorieWithModelesSerializer(data=request.data)
#         if not serializer.is_valid():
#             return Response(serializer.errors, status=400)

#         nom_categorie = serializer.validated_data["nom"].strip().title()
#         modeles_noms = serializer.validated_data["modeles"]

#         try:
#             categorie = Categorie.objects.get(nom=nom_categorie)
#         except Categorie.DoesNotExist:
#             return Response({"error": "Catégorie introuvable."}, status=404)

#         nouveaux = []
#         deja_existants = []

#         for nom in modeles_noms:
#             modele, _ = Modele.objects.get_or_create(modele=nom)
#             obj, created = CategorieModele.objects.get_or_create(categorie=categorie, modele=modele)
#             if created:
#                 nouveaux.append(modele.modele)
#             else:
#                 deja_existants.append(modele.modele)

#         return Response({
#             "message": "🔄 Mise à jour effectuée.",
#             "categorie": categorie.nom,
#             "ajoutes": nouveaux,
#             "deja_existants": deja_existants
#         }, status=200)


class ModeleUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Modifier le nom d’un modèle existant",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["id", "nouveau_nom"],
            properties={
                "id": openapi.Schema(type=openapi.TYPE_INTEGER, description="ID du modèle à modifier"),
                "nouveau_nom": openapi.Schema(type=openapi.TYPE_STRING, description="Nouveau nom du modèle"),
            }
        ),
        responses={200: "Modèle mis à jour", 404: "Modèle non trouvé"}
    )
    def patch(self, request):
        modele_id = request.data.get("id")
        nouveau_nom = request.data.get("nouveau_nom")

        try:
            modele = Modele.objects.get(id=modele_id)
        except Modele.DoesNotExist:
            return Response({"error": "Modèle introuvable."}, status=404)

        ancien_nom = modele.modele
        modele.modele = nouveau_nom.strip().title()
        modele.save()

        return Response({
            "message": "✅ Modèle mis à jour.",
            "ancien_nom": ancien_nom,
            "nouveau_nom": modele.modele,
        }, status=200)


# class MarqueCreateWithCategorieView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Créer une marque avec liaison à des catégories",
#         operation_description="Crée une nouvelle marque et lie celle-ci à une ou plusieurs catégories via la table CategorieMarque.",
#         request_body=openapi.Schema(
#             type=openapi.TYPE_OBJECT,
#             required=["marque", "purete", "prix", "categories"],
#             properties={
#                 "marque": openapi.Schema(type=openapi.TYPE_STRING, description="Nom de la marque"),
#                 "purete": openapi.Schema(type=openapi.TYPE_INTEGER, description="ID de la pureté"),
#                 "prix": openapi.Schema(type=openapi.TYPE_NUMBER, format="decimal", description="Prix de base de la marque"),
#                 "categories": openapi.Schema(
#                     type=openapi.TYPE_ARRAY,
#                     items=openapi.Items(type=openapi.TYPE_INTEGER),
#                     description="Liste des IDs de catégories"
#                 )
#             }
#         ),
#         responses={
#             201: openapi.Response(description="Marque créée avec succès."),
#             400: "Requête invalide"
#         }
#     )
#     def post(self, request):
#         data = request.data
#         marque_nom = data.get("marque", "").strip()
#         purete_id = data.get("purete")
#         prix = data.get("prix")
#         categories_ids = data.get("categories", [])

#         # Validation de base
#         if not marque_nom:
#             return Response({"error": "Le nom de la marque est requis."}, status=400)

#         try:
#             prix = Decimal(str(prix))
#             if prix < 0:
#                 raise InvalidOperation
#         except (InvalidOperation, TypeError):
#             return Response({"error": "Prix invalide."}, status=400)

#         # Vérification de la pureté
#         try:
#             purete = Purete.objects.get(id=purete_id)
#         except Purete.DoesNotExist:
#             return Response({"error": "Pureté non trouvée."}, status=404)

#         # Création de la marque
#         marque = Marque.objects.create(
#             marque=marque_nom,
#             purete=purete,
#             prix=prix
#         )

#         # Liaison aux catégories
#         categories_liees = []
#         for cat_id in categories_ids:
#             try:
#                 categorie = Categorie.objects.get(id=cat_id)
#                 CategorieMarque.objects.get_or_create(categorie=categorie, marque=marque)
#                 categories_liees.append(categorie.nom)
#             except Categorie.DoesNotExist:
#                 continue

#         return Response({
#             "message": "✅ Marque créée avec succès.",
#             "marque": marque.marque,
#             "prix": str(marque.prix),
#             "purete": str(marque.purete),
#             "categories_liees": categories_liees
#         }, status=201)


# class MarqueUpdateWithCategoriesView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Mettre à jour une marque et ses catégories associées",
#         operation_description="""
#         Met à jour le nom, la pureté, le prix d'une marque existante et remplace ses catégories associées.
#         """,
#         manual_parameters=[
#             openapi.Parameter(
#                 'marque_id',
#                 openapi.IN_PATH,
#                 description="ID de la marque à mettre à jour",
#                 type=openapi.TYPE_INTEGER,
#                 required=True
#             )
#         ],
#         request_body=openapi.Schema(
#             type=openapi.TYPE_OBJECT,
#             required=["marque", "purete", "prix", "categories"],
#             properties={
#                 "marque": openapi.Schema(type=openapi.TYPE_STRING),
#                 "purete": openapi.Schema(type=openapi.TYPE_INTEGER),
#                 "prix": openapi.Schema(type=openapi.TYPE_NUMBER, format="decimal"),
#                 "categories": openapi.Schema(
#                     type=openapi.TYPE_ARRAY,
#                     items=openapi.Items(type=openapi.TYPE_INTEGER),
#                     description="Liste des IDs des nouvelles catégories"
#                 )
#             }
#         ),
#         responses={
#             200: openapi.Response(description="Marque mise à jour avec succès"),
#             404: "Marque ou pureté introuvable",
#             400: "Requête invalide"
#         }
#     )
#     def put(self, request, marque_id):
#         data = request.data
#         nom = data.get("marque", "").strip()
#         purete_id = data.get("purete")
#         prix = data.get("prix")
#         categories_ids = data.get("categories", [])

#         try:
#             marque = Marque.objects.get(id=marque_id)
#         except Marque.DoesNotExist:
#             return Response({"error": "Marque introuvable."}, status=404)

#         try:
#             purete = Purete.objects.get(id=purete_id)
#         except Purete.DoesNotExist:
#             return Response({"error": "Pureté introuvable."}, status=404)

#         try:
#             prix = Decimal(str(prix))
#         except (InvalidOperation, TypeError):
#             return Response({"error": "Prix invalide."}, status=400)

#         marque.marque = nom or marque.marque
#         marque.purete = purete
#         marque.prix = prix
#         marque.save()

#         # ✅ On supprime les anciennes liaisons
#         CategorieMarque.objects.filter(marque=marque).delete()

#         # ✅ On ajoute les nouvelles liaisons
#         categories_liees = []
#         for cat_id in categories_ids:
#             try:
#                 categorie = Categorie.objects.get(id=cat_id)
#                 CategorieMarque.objects.create(categorie=categorie, marque=marque)
#                 categories_liees.append(categorie.nom)
#             except Categorie.DoesNotExist:
#                 continue

#         return Response({
#             "message": "✅ Marque mise à jour avec succès.",
#             "marque": marque.marque,
#             "prix": str(marque.prix),
#             "purete": str(marque.purete),
#             "categories_liees": categories_liees
#         }, status=200)



# # Suppression propre des liens CategorieMarque (sans supprimer la Marque)
# class SupprimerCategoriesDeMarqueView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Supprimer toutes les catégories liées à une marque",
#         manual_parameters=[
#             openapi.Parameter('marque_id', openapi.IN_PATH, description="ID de la marque", type=openapi.TYPE_INTEGER)
#         ],
#         responses={200: "✅ Liens supprimés", 404: "Marque introuvable"}
#     )
#     def delete(self, request, marque_id):
#         try:
#             marque = Marque.objects.get(id=marque_id)
#         except Marque.DoesNotExist:
#             return Response({"error": "Marque introuvable."}, status=404)

#         nb_deleted, _ = CategorieMarque.objects.filter(marque=marque).delete()

#         return Response({
#             "message": f"✅ {nb_deleted} lien(s) supprimé(s) entre la marque et ses catégories."
#         })


class MarquePartialUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Mise à jour partielle d'une marque (PATCH)",
        operation_description="""
        Permet de mettre à jour certains champs de la marque (nom, prix, pureté) et/ou ses catégories associées.
        L’envoi du champ `categories` (liste d’IDs) écrase les liaisons existantes.
        """,
        manual_parameters=[
            openapi.Parameter('marque_id', openapi.IN_PATH, description="ID de la marque", type=openapi.TYPE_INTEGER)
        ],
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "marque": openapi.Schema(type=openapi.TYPE_STRING),
                "prix": openapi.Schema(type=openapi.TYPE_NUMBER, format="decimal"),
                "purete": openapi.Schema(type=openapi.TYPE_INTEGER),
                "categories": openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Items(type=openapi.TYPE_INTEGER),
                    description="Remplace les liaisons actuelles (optionnel)"
                )
            }
        ),
        responses={200: "✅ Mise à jour partielle réussie", 404: "Marque ou pureté introuvable"}
    )
    def patch(self, request, marque_id):
        data = request.data
        try:
            marque = Marque.objects.get(id=marque_id)
        except Marque.DoesNotExist:
            return Response({"error": "Marque introuvable."}, status=404)

        if 'marque' in data:
            marque.marque = data['marque'].strip()

        if 'prix' in data:
            try:
                marque.prix = Decimal(str(data['prix']))
            except (InvalidOperation, TypeError):
                return Response({"error": "Prix invalide."}, status=400)

        if 'purete' in data:
            try:
                marque.purete = Purete.objects.get(id=data['purete'])
            except Purete.DoesNotExist:
                return Response({"error": "Pureté introuvable."}, status=404)

        return Response({"message": "✅ Mise à jour partielle réussie."})



# class ModeleListAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]

#     # ✅ Rôles autorisés
#     allowed_roles_admin_manager = ['admin', 'manager']

#     @swagger_auto_schema(
#         operation_description="Lister tous les modèles, avec possibilité de filtrer par nom ou catégorie.",
#         manual_parameters=[
#             openapi.Parameter(
#                 'nom', openapi.IN_QUERY,
#                 description="Nom du modèle (recherche partielle)",
#                 type=openapi.TYPE_STRING
#             ),
#             openapi.Parameter(
#                 'marque_id', openapi.IN_QUERY,
#                 description="ID de la marque",
#                 type=openapi.TYPE_INTEGER
#             )
#         ],
#         responses={
#             200: openapi.Response("Liste des modèles", ModeleSerializer(many=True)),
#             403: "⛔ Accès refusé"
#         }
#     )
#     def get(self, request):
#         user = request.user
#         if not user.user_role or user.user_role.role not in self.allowed_roles_admin_manager:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         queryset = Modele.objects.all()
#         nom = request.GET.get('nom')
#         marque_id = request.GET.get('marque_id')

#         if nom:
#             queryset = queryset.filter(modele__icontains=nom)
#         if marque_id:
#             queryset = queryset.filter(marque_id=marque_id)

#         serializer = ModeleSerializer(queryset, many=True)
#         return Response(serializer.data)


# class ModeleCategorieListAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les modèles avec leurs catégories associées",
#         responses={200: openapi.Response("Liste des modèles avec catégories", ModeleWithCategoriesSerializer(many=True))}
#     )
#     def get(self, request):
#         modeles = Modele.objects.all().prefetch_related('modele_categories__categorie')
#         serializer = ModeleWithCategoriesSerializer(modeles, many=True)
#         return Response(serializer.data)


class ModeleListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Lister les modèles avec leurs catégories associées",
        responses={200: ModeleListSerializer(many=True)}
    )
    def get(self, request):
        modeles = Modele.objects.prefetch_related("modele_categories__categorie").all()
        serializer = ModeleListSerializer(modeles, many=True)
        return Response(serializer.data)


# class ModeleListAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]

#     allowed_roles_admin_manager = ['admin', 'manager']

#     @swagger_auto_schema(
#         operation_description="Lister tous les modèles avec filtre par nom, catégorie et marque.",
#         manual_parameters=[
#             openapi.Parameter('nom', openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Recherche par nom de modèle"),
#             openapi.Parameter('marque_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description="ID de la marque"),
#         ],
#         responses={
#             200: openapi.Response("Liste des modèles", ModeleSerializer(many=True)),
#             403: "⛔ Accès refusé"
#         }
#     )
#     def get(self, request):
#         user = request.user
#         if not user.user_role or user.user_role.role not in self.allowed_roles_admin_manager:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         queryset = Modele.objects.all()
#         nom = request.GET.get('nom')
#         categorie_id = request.GET.get('categorie_id')
#         marque_id = request.GET.get('marque_id')  # 🆕

#         if nom:
#             queryset = queryset.filter(modele__icontains=nom)
#         if categorie_id:
#             queryset = queryset.filter(categorie_id=categorie_id)
#         if marque_id:
#             queryset = queryset.filter(marque_id=marque_id)

#         serializer = ModeleSerializer(queryset, many=True)
#         return Response(serializer.data)


class ModeleCreateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    # ✅ Rôles autorisés pour créer un modèle
    allowed_roles_admin_manager = ['admin', 'manager']

    @swagger_auto_schema(
        operation_description="Créer un nouveau modèle en utilisant le nom de la catégorie.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["modele", "marque"],
            properties={
                "modele": openapi.Schema(type=openapi.TYPE_STRING, description="Nom du modèle"),
                "marque": openapi.Schema(type=openapi.TYPE_STRING, description="Nom de la marque (ex: 'local')"),
            },
            example={
                "modele": "Alliance homme or jaune",
                "marque": "local"
            }
        ),
        responses={
            201: openapi.Response("Modèle créé avec succès", ModeleSerializer),
            400: "Erreur de validation",
            403: "⛔ Accès refusé"
        }
    )
    def post(self, request):
        user = request.user
        if not user.user_role or user.user_role.role not in self.allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        serializer = ModeleSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ModeleUpdateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    # ✅ Liste des rôles autorisés
    allowed_roles_admin_manager = ['admin', 'manager']

    def get_object(self, pk):
        try:
            return Modele.objects.get(pk=pk)
        except Modele.DoesNotExist:
            return None

    @swagger_auto_schema(
        operation_description="🛠 Modifier complètement un modèle (PUT)",
        request_body=ModeleSerializer,
        responses={
            200: openapi.Response("Modèle mis à jour avec succès", ModeleSerializer),
            400: "Requête invalide",
            403: "Accès refusé",
            404: "Modèle introuvable"
        }
    )
    def put(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in self.allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        modele_instance = self.get_object(pk)
        if modele_instance is None:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = ModeleSerializer(modele_instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(
        operation_description="✏️ Modifier partiellement un modèle (PATCH)",
        request_body=ModeleSerializer,
        responses={
            200: openapi.Response("Modèle mis à jour partiellement", ModeleSerializer),
            400: "Requête invalide",
            403: "Accès refusé",
            404: "Modèle introuvable"
        }
    )
    def patch(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in self.allowed_roles_admin_manager:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        modele = self.get_object(pk)
        if modele is None:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = ModeleSerializer(modele, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class ModeleDeleteAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    def get_object(self, pk):
        try:
            return Modele.objects.get(pk=pk)
        except Modele.DoesNotExist:
            return None

    def delete(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        
        modele_instance = self.get_object(pk)
        if modele_instance is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        modele_instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# a. Lister les purete par catégorie
# class PureteParCategorieAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les puretés d'une catégorie",
#         operation_description="Retourne toutes les puretés associées à une catégorie spécifiée par son nom (insensible à la casse).",
#         manual_parameters=[
#             openapi.Parameter(
#                 'categorie', openapi.IN_QUERY,
#                 description="Nom exact de la catégorie (insensible à la casse)",
#                 type=openapi.TYPE_STRING,
#                 required=True
#             ),
#         ],
#         responses={200: openapi.Response("Liste des puretés", PureteSerializer(many=True))}
#     )
#     def get(self, request):
#         nom_categorie = request.GET.get('categorie')
#         if not nom_categorie:
#             return Response({"error": "Le paramètre 'categorie' est requis."}, status=400)

#         try:
#             categorie = Categorie.objects.get(nom__iexact=nom_categorie.strip())
#         except Categorie.DoesNotExist:
#             return Response({"error": "Catégorie non trouvée."}, status=404)

#         puretes = Purete.objects.filter(categorie=categorie)
#         serializer = PureteSerializer(puretes, many=True)
#         return Response(serializer.data)

# b. Lister les marque par purete
# class MarqueParCategorieAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les marques selon le nom du categorie",
#         manual_parameters=[
#             openapi.Parameter(
#                 'categorie', openapi.IN_QUERY,
#                 description="Nom exact de categorie (ex: bague)",
#                 type=openapi.TYPE_STRING,
#                 required=True
#             ),
#         ],
#         responses={200: openapi.Response("Liste des marques", MarqueSerializer(many=True))}
#     )
#     def get(self, request):
#         nom_categorie = request.GET.get('categorie')
#         if not nom_categorie:
#             return Response({"error": "Le paramètre 'categorie' est requis."}, status=400)

#         try:
#             categorie = Categorie.objects.get(categorie__iexact=nom_categorie)
#         except Categorie.DoesNotExist:
#             return Response({"error": "Categorie non trouvée."}, status=404)

#         marques = Marque.objects.filter(categorie=categorie)
#         serializer = MarqueSerializer(marques, many=True)
#         return Response(serializer.data)

# class MarqueParCategorieAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les marques selon le nom de la catégorie",
#         manual_parameters=[
#             openapi.Parameter(
#                 'categorie', openapi.IN_QUERY,
#                 description="Nom exact de la catégorie (ex: bague)",
#                 type=openapi.TYPE_STRING,
#                 required=True
#             ),
#         ],
#         responses={200: openapi.Response("Liste des marques", MarqueSerializer(many=True))}
#     )
#     def get(self, request):
#         nom_categorie = request.GET.get('categorie')
#         if not nom_categorie:
#             return Response({"error": "Le paramètre 'categorie' est requis."}, status=400)

#         try:
#             categorie = Categorie.objects.get(nom__iexact=nom_categorie)
#         except Categorie.DoesNotExist:
#             return Response({"error": "Catégorie non trouvée."}, status=404)

#         marques = Marque.objects.filter(categorie=categorie)
#         serializer = MarqueSerializer(marques, many=True)
#         return Response(serializer.data)

# b. Lister les modèles par marque
# class ModeleParMarqueAPIView(APIView):
#     @swagger_auto_schema(
#         operation_summary="Lister les modèles d'une marque (par nom)",
#         manual_parameters=[
#             openapi.Parameter(
#                 'marque', openapi.IN_QUERY,
#                 description="Nom exact de la marque",
#                 type=openapi.TYPE_STRING,
#                 required=True
#             ),
#         ],
#         responses={200: openapi.Response("Liste des modèles", ModeleSerializer(many=True))}
#     )
#     def get(self, request):
#         nom_marque = request.GET.get('marque')
#         if not nom_marque:
#             return Response({"error": "Le paramètre 'marque' est requis."}, status=400)

#         try:
#             marque = Marque.objects.get(marque__iexact=nom_marque)
#         except Marque.DoesNotExist:
#             return Response({"error": "Marque non trouvée."}, status=404)

#         modeles = Modele.objects.filter(marque=marque)
#         serializer = ModeleSerializer(modeles, many=True)
#         return Response(serializer.data)


class ProduitListAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Liste tous les rôles disponibles",
        responses={200: ProduitSerializer(many=True)},
        manual_parameters=[
            openapi.Parameter(
                'search',
                openapi.IN_QUERY,
                description="Filtrer les rôles par sku",
                type=openapi.TYPE_STRING
            )
        ]
    )
    def get(self, request):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager', 'vendor']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        
        search = request.GET.get('search')
        queryset = Produit.objects.all()
        if search:
            queryset = queryset.filter(sku__icontains=search)
        serializer = ProduitSerializer(queryset, many=True)
        return Response(serializer.data)


# class ProduitCreateAPIView(APIView):
#     parser_classes = [MultiPartParser, FormParser]
#     permission_classes = [IsAuthenticated]
#     renderer_classes = [UserRenderer]

#     @swagger_auto_schema(
#         operation_summary="Créer un produit avec images et QR code",
#         operation_description="Crée un produit en utilisant les noms de la catégorie, pureté, marque et modèle.",
#         manual_parameters=[
#             openapi.Parameter('nom', openapi.IN_FORM, type=openapi.TYPE_STRING),
#             openapi.Parameter('image', openapi.IN_FORM, type=openapi.TYPE_FILE),
#             openapi.Parameter('description', openapi.IN_FORM, type=openapi.TYPE_STRING),
#             openapi.Parameter('genre', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['F', 'H', 'E'], default='F'),
#             openapi.Parameter('categorie', openapi.IN_FORM, type=openapi.TYPE_STRING, description="Nom de la catégorie"),
#             openapi.Parameter('purete', openapi.IN_FORM, type=openapi.TYPE_STRING, description="Ex: 18"),
#             openapi.Parameter('marque', openapi.IN_FORM, type=openapi.TYPE_STRING, description="Nom de la marque"),
#             openapi.Parameter('modele', openapi.IN_FORM, type=openapi.TYPE_STRING, description="Nom du modèle"),
#             openapi.Parameter('matiere', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['or', 'argent', 'mixte'], default='or'),
#             openapi.Parameter('poids', openapi.IN_FORM, type=openapi.TYPE_NUMBER),
#             openapi.Parameter('taille', openapi.IN_FORM, type=openapi.TYPE_NUMBER),
#             openapi.Parameter('status', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['publié', 'désactivé', 'rejetée'], default='publié'),
#             openapi.Parameter('etat', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['N', 'R'], default='N'),
#             openapi.Parameter('gallery', openapi.IN_FORM, type=openapi.TYPE_FILE, description="Fichiers galerie", required=False, multiple=True),
#         ],
#         responses={
#             201: openapi.Response("Produit créé avec succès", ProduitSerializer),
#             400: "Erreur de validation"
#         }
#     )
#     @transaction.atomic
#     def post(self, request):
#         user = request.user
#         if not user.user_role or user.user_role.role not in ['admin', 'manager']:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         data = request.data.copy()

#         # Remplacement des noms par des IDs
#         try:
#             data['categorie'] = Categorie.objects.get(nom__iexact=data.get('categorie')).id
#         except Categorie.DoesNotExist:
#             return Response({"error": "Catégorie introuvable"}, status=400)

#         try:
#             data['purete'] = Purete.objects.get(purete__iexact=data.get('purete')).id
#         except Purete.DoesNotExist:
#             return Response({"error": "Pureté introuvable"}, status=400)

#         try:
#             data['marque'] = Marque.objects.get(marque__iexact=data.get('marque')).id
#         except Marque.DoesNotExist:
#             return Response({"error": "Marque introuvable"}, status=400)

#         try:
#             data['modele'] = Modele.objects.get(modele__iexact=data.get('modele')).id
#         except Modele.DoesNotExist:
#             return Response({"error": "Modèle introuvable"}, status=400)

#         # Création
#         serializer = ProduitSerializer(data=data, context={"request": request})
#         if serializer.is_valid():
#             produit = serializer.save()

#             for image in request.FILES.getlist("gallery"):
#                 Gallery.objects.create(produit=produit, image=image)

#             produit.refresh_from_db()
#             return Response(ProduitSerializer(produit, context={"request": request}).data, status=201)

#         return Response(serializer.errors, status=400)






def get_object_id_or_error(model, field, value, label):
    try:
        return model.objects.get(**{f"{field}__iexact": value}).id
    except model.DoesNotExist:
        raise ValidationError({label: f"{label.capitalize()} '{value}' introuvable."})

class ProduitCreateAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]
    renderer_classes = [UserRenderer]

    @swagger_auto_schema(
        operation_summary="Créer un produit avec images et QR code",
        operation_description="Crée un produit en utilisant les noms de la catégorie, pureté, marque et modèle.",
        manual_parameters=[
            openapi.Parameter('nom', openapi.IN_FORM, type=openapi.TYPE_STRING),
            openapi.Parameter('image', openapi.IN_FORM, type=openapi.TYPE_FILE),
            openapi.Parameter('description', openapi.IN_FORM, type=openapi.TYPE_STRING),
            openapi.Parameter('genre', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['F', 'H', 'E'], default='F'),
            openapi.Parameter('categorie', openapi.IN_FORM, type=openapi.TYPE_STRING),
            openapi.Parameter('purete', openapi.IN_FORM, type=openapi.TYPE_STRING),
            openapi.Parameter('marque', openapi.IN_FORM, type=openapi.TYPE_STRING),
            openapi.Parameter('modele', openapi.IN_FORM, type=openapi.TYPE_STRING),
            openapi.Parameter('matiere', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['or', 'argent', 'mixte'], default='or'),
            openapi.Parameter('poids', openapi.IN_FORM, type=openapi.TYPE_NUMBER),
            openapi.Parameter('taille', openapi.IN_FORM, type=openapi.TYPE_NUMBER),
            openapi.Parameter('status', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['publié', 'désactivé', 'rejetée'], default='publié'),
            openapi.Parameter('etat', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['N', 'R'], default='N'),
            openapi.Parameter('gallery', openapi.IN_FORM, type=openapi.TYPE_FILE, required=False, multiple=True),
        ],
        responses={
            201: openapi.Response("Produit créé avec succès", ProduitSerializer),
            400: "Erreur de validation"
        }
    )
    @transaction.atomic
    def post(self, request):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

        data = request.data.copy()

        try:
            
            # Création du produit
            serializer = ProduitSerializer(data=data, context={"request": request})
            if serializer.is_valid():
                produit = serializer.save()

                # Ajout des images
                images = request.FILES.getlist("gallery")
                for image in images:
                    Gallery.objects.create(produit=produit, image=image)

                produit.refresh_from_db()

                return Response({
                    "message": f"Produit créé avec succès avec {len(images)} image(s)",
                    "produit": ProduitSerializer(produit, context={"request": request}).data
                }, status=201)
            else:
                return Response(serializer.errors, status=400)

        except Categorie.DoesNotExist:
            return Response({"error": "Catégorie introuvable"}, status=400)
        except Purete.DoesNotExist:
            return Response({"error": "Pureté introuvable"}, status=400)
        except Marque.DoesNotExist:
            return Response({"error": "Marque introuvable"}, status=400)
        except Modele.DoesNotExist:
            return Response({"error": "Modèle introuvable"}, status=400)
        except Exception as e:
            return Response({"error": f"Une erreur est survenue : {str(e)}"}, status=500)


# class ProduitCreateAPIView(APIView):
#     parser_classes = [MultiPartParser, FormParser]
#     permission_classes = [IsAuthenticated]
#     renderer_classes = [UserRenderer]

#     @swagger_auto_schema(
#         operation_summary="Créer un produit avec images et QR code",
#         manual_parameters=[
#             openapi.Parameter('image', openapi.IN_FORM, type=openapi.TYPE_FILE),
#             openapi.Parameter('nom', openapi.IN_FORM, type=openapi.TYPE_STRING),
#             openapi.Parameter('description', openapi.IN_FORM, type=openapi.TYPE_STRING),
#             openapi.Parameter('genre', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['F', 'H', 'E'], default='F'),
#             openapi.Parameter('categorie', openapi.IN_FORM, type=openapi.TYPE_INTEGER, description="ID de la catégorie"),
#             openapi.Parameter('marque', openapi.IN_FORM, type=openapi.TYPE_INTEGER, description="ID de la marque"),
#             openapi.Parameter('modele', openapi.IN_FORM, type=openapi.TYPE_INTEGER, description="ID du modèle"),
#             openapi.Parameter('purete', openapi.IN_FORM, type=openapi.TYPE_INTEGER, description="ID de la pureté"),
#             openapi.Parameter('matiere', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['or', 'argent', 'mixte'], default='or'),
#             openapi.Parameter('poids', openapi.IN_FORM, type=openapi.TYPE_NUMBER),
#             openapi.Parameter('taille', openapi.IN_FORM, type=openapi.TYPE_NUMBER),
#             openapi.Parameter('status', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['publié', 'désactivé', 'rejetée'], default='publié'),
#             openapi.Parameter('etat', openapi.IN_FORM, type=openapi.TYPE_STRING, enum=['N', 'R'], default='N'),
#             openapi.Parameter('gallery', openapi.IN_FORM, type=openapi.TYPE_FILE, description="Fichiers galerie", required=False, multiple=True),
#         ],
#         responses={
#             201: openapi.Response("Produit créé", ProduitSerializer),
#             400: "Erreur de validation"
#         }
#     )
#     @transaction.atomic
#     def post(self, request):
#         user = request.user
#         if not user.user_role or user.user_role.role not in ['admin', 'manager']:
#             return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)

#         try:
#             data = request.data.copy()

#             # Crée le produit avec les données POST
#             serializer = ProduitSerializer(data=data, context={"request": request})
#             if serializer.is_valid():
#                 produit = serializer.save()

#                 # Ajouter les images dans la galerie
#                 images = request.FILES.getlist("gallery")
#                 for image in images:
#                     Gallery.objects.create(produit=produit, image=image)

#                 produit.refresh_from_db()
#                 return Response({
#                     "message": f"Produit créé avec succès avec {len(images)} image(s)",
#                     "produit": ProduitSerializer(produit, context={"request": request}).data
#                 }, status=201)
#             else:
#                 return Response(serializer.errors, status=400)

#         except Exception as e:
#             return Response({"error": f"Une erreur est survenue : {str(e)}"}, status=500)
        

class ProduitDetailSlugView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="🧾 Détail d’un produit via son slug",
        operation_description="Retourne les informations complètes d’un produit en le récupérant par son `slug`.",
        responses={
            200: openapi.Response("Détail du produit", ProduitSerializer),
            404: "Produit non trouvé"
        },
        manual_parameters=[
            openapi.Parameter(
                'slug',
                openapi.IN_PATH,
                description="Slug du produit (ex: bague-or-abc123)",
                type=openapi.TYPE_STRING,
                required=True
            )
        ]
    )
    def get(self, request, slug):
        try:
            produit = Produit.objects.get(slug=slug)
            serializer = ProduitSerializer(produit, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Produit.DoesNotExist:
            return Response({"error": "Produit non trouvé."}, status=status.HTTP_404_NOT_FOUND)




# class ProduitCreateAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]
    
#     @swagger_auto_schema(
#         operation_description="Créer un produit avec sa galerie d’images.",
#         request_body=ProduitSerializer,
#         responses={
#             status.HTTP_201_CREATED: openapi.Response('User created successfully', ProduitSerializer),
#             status.HTTP_400_BAD_REQUEST: openapi.Response('Bad Request')
#         }
#     )
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         try:
#             # 1. Désérialisation des données produit
#             produit_data = {
#                 'nom': request.data.get('nom'),
#                 'image': request.data.get('image'),
#                 'genre': request.data.get('genre'),
#                 'categorie': request.data.get('categorie'),
#                 'marque': request.data.get('marque'),
#                 'modele': request.data.get('modele'),
#                 'purete': request.data.get('purete'),
#                 'matiere': request.data.get('matiere'),
#                 'poids': request.data.get('poids'),
#                 'taille': request.data.get('taille'),
#                 'status': request.data.get('status'),
#             }
#             produit_serializer = ProduitSerializer(data=produit_data)
#             if not produit_serializer.is_valid():
#                 return Response(produit_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#             produit = produit_serializer.save()
#             # 2. Désérialisation des fichiers image
#             images = request.FILES.getlist('gallery')  # récupère tous les fichiers avec le champ "gallery"
#             for image_file in images:
#                 Gallery.objects.create(produit=produit, image=image_file)
#             # return Response({
#             #     'message': 'Produit et galerie créés avec succès',
#             #     'produit_id': produit.id
#             # }, status=status.HTTP_201_CREATED)
#             return Response(produit_serializer.data, status=status.HTTP_201_CREATED)
#         except Exception as e:
#             return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
    
    # def post(self, request):
    #     if request.user.user_role is not None and request.user.user_role.role != 'admin' and request.user.user_role.role != 'manager':
    #         return Response({"message": "Access Denied"})
    #     # if request.user.is_authenticated and request.user.user_role and not request.user.user_role.role == 'admin' and not request.user.user_role.role == 'manager' and not request.user.user_role.role == 'seller':
    #     #     return Response({"message": "Access Denied"})
    #     serializer = ProduitSerializer(data=request.data)
    #     if serializer.is_valid():
    #         produit = serializer.save()
    #         # produit.save()
    #         return Response(serializer.data, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    # @transaction.atomic
    # def post(self, request, *args, **kwargs):
    #     # First, deserialize the input data for the produit and stock
    #     produit_data = request.data.get('produit')
    #     stock_data = request.data.get('stock')

    #     # Validate and create the produit
    #     produit_serializer = ProduitSerializer(data=produit_data)
    #     if produit_serializer.is_valid():
    #         produit = produit_serializer.save()  # This will save the produit to the database
    #     else:
    #         return Response(produit_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    #     # Validate and create the stock
    #     stock_data['produit'] = produit.id  # Associate the stock with the created produit
    #     stock_serializer = StockSerializer(data=stock_data)
    #     if stock_serializer.is_valid():
    #         stock_serializer.save()  # This will save the stock to the database
    #     else:
    #         return Response(stock_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    #     return Response({
    #         'produit': produit_serializer.data,
    #         'stock': stock_serializer.data
    #     }, status=status.HTTP_201_CREATED)


class ProduitGetOneAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    
    def get_object(self, pk):
        try:
            return Produit.objects.get(pk=pk)
        except Produit.DoesNotExist:
            return None

    @swagger_auto_schema(
        operation_description="Récupère un produit par ID",
        responses={200: ProduitSerializer, 404: "Produit non trouvé"}
    )
    def get(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        
        produit = self.get_object(pk)
        if produit is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = ProduitSerializer(produit)
        return Response(serializer.data)


class ProduitUpdateAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    
    def get_object(self, pk):
        try:
            return Produit.objects.get(pk=pk)
        except Produit.DoesNotExist:
            raise Http404
        
    @swagger_auto_schema(
        operation_description="Mise à jour complète d'un produit",
        request_body=ProduitSerializer,
        responses={
            200: ProduitSerializer,
            400: "Requête invalide",
            404: "Produit non trouvé"
        }
    )
    def put(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        
        # if request.user.is_authenticated and request.user.user_role and not request.user.user_role.role == 'admin' and not request.user.user_role.role == 'manager' and not request.user.user_role.role == 'seller':
        #     return Response({"message": "Access Denied"})
        produit = self.get_object(pk)
        if produit is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = ProduitSerializer(produit, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @swagger_auto_schema(
        operation_description="Mise à jour partielle d'un produit",
        request_body=ProduitSerializer,
        responses={
            200: ProduitSerializer,
            400: "Requête invalide",
            404: "Rôle non trouvé"
        }
    )
    def patch(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        produit = self.get_object(pk)
        serializer = ProduitSerializer(produit, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProduitDeleteAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    
    def get_object(self, pk):
        try:
            return Produit.objects.get(pk=pk)
        except Produit.DoesNotExist:
            return None
    @swagger_auto_schema(
        operation_description="Supprime un produit par ID",
        responses={
            204: 'Supprimé avec succès',
            404: 'Produit non trouvé'
        }
    )
    def delete(self, request, pk):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        
        produit = self.get_object(pk)
        if produit is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        produit.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)



class QRCodeView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        responses={200: openapi.Response('response description', ProduitSerializer)},
        )
    def get(self, request, pk, format=None):
        user = request.user
        if not user.user_role or user.user_role.role not in ['admin', 'manager']:
            return Response({"message": "Access Denied"}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            # Retrieve the produit by its ID
            produit = Produit.objects.get(pk=pk)
            # Generate QR code data, e.g., produit URL or information
            # qr_data = f"Produit Nom: {produit.categorie.nom} {produit.modele} {produit.marque} {produit.purete}\nPrix gramme: {produit.marque.prix}\nDescription: {produit.description}"
            qr_data = f"Produit qr-code: {produit.slug}"
            # Create a QR code
            qr = qrcode.make(qr_data)
            # Save the QR code in a BytesIO object
            img_io = BytesIO()
            qr.save(img_io)
            img_io.seek(0)
            # Return the image as an HTTP response
            return HttpResponse(img_io, content_type="image/png")
        
        except Produit.DoesNotExist:
            return Response({"error": "Produit not found"}, status=status.HTTP_404_NOT_FOUND)
        


# class ProduitQRCodeView(APIView):
#     def get(self, request, pk):
#         produit = get_object_or_404(Produit, pk=pk)
#         data = f"Produit: {produit.name}\nDescription: {produit.description}"
        
#         qr = qrcode.make(data)
#         buffer = BytesIO()
#         qr.save(buffer, format="PNG")
#         buffer.seek(0)

#         return HttpResponse(buffer, content_type="image/png")


# #list qr_code in excel
# class ExportQRCodeExcelAPIView(APIView):
#     renderer_classes = [UserRenderer]
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         responses={200: openapi.Response('response description', ProduitSerializer)},
#         )
#     def get(self, request):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "QR Codes Produits"

#         # En-têtes
#         ws.append(["Nom du produit", "QR Code"])

#         produits = Produit.objects.all()
#         row = 2

#         for produit in produits:
#             # Générer le QR code
#             data = f"Produit ID: {produit.id}, Nom: {produit.nom}"
#             qr = qrcode.make(data)

#             # Sauvegarde temporaire de l’image
#             buffer = BytesIO()
#             qr.save(buffer, format="PNG")
#             buffer.seek(0)

#             # Créer image PIL compatible openpyxl
#             img = XLImage(buffer)
#             img.width = 50
#             img.height = 50

#             # Insérer le nom et l’image QR code
#             ws.cell(row=row, column=1, value=produit.nom)
#             cell = f'B{row}'
#             ws.add_image(img, cell)

#             row += 1

#         # Export Excel dans un buffer
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename="qr_codes_produits.xlsx"'

#         excel_buffer = BytesIO()
#         wb.save(excel_buffer)
#         response.write(excel_buffer.getvalue())

#         return response


class ExportOneQRCodeExcelAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        responses={200: openapi.Response('response description', ProduitSerializer)},
        )
    def get(self, request, slug):
        try:
            produit = Produit.objects.get(slug=slug)
        except Produit.DoesNotExist:
            raise Http404("Produit non trouvé")

        # Créer le QR code
        data = f"Produit SKU: {produit.slug}"
        qr = qrcode.make(data)

        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        buffer.seek(0)

        # Préparer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "QR Code Produit"
        ws.append(["Nom du produit", "QR Code"])
        ws.cell(row=2, column=1, value=produit.slug)

        # Ajouter l’image
        img = XLImage(buffer)
        img.width = 100
        img.height = 100
        ws.add_image(img, "B2")

        # Export
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"qr_code_produit_{produit.slug}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        response.write(excel_buffer.getvalue())

        return response



# class GetGalleryByProduitAPIView(APIView):
#     permission_classes = [IsAuthenticated]

#     @swagger_auto_schema(
#         operation_summary="Lister les images de la galerie d’un produit",
#         manual_parameters=[
#             openapi.Parameter('produit_id', openapi.IN_QUERY, description="ID du produit", type=openapi.TYPE_INTEGER)
#         ],
#         responses={200: GallerySerializer(many=True)}
#     )
#     def get(self, request):
#         produit_id = request.GET.get('produit_id')
#         if not produit_id:
#             return Response({"error": "produit_id requis"}, status=400)

#         galleries = Gallery.objects.filter(produit_id=produit_id)
#         serializer = GallerySerializer(galleries, many=True)
#         return Response(serializer.data)


class ProduitRecentListAPIView(APIView):
    # permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Lister les produits actifs les plus récents",
        responses={200: ProduitWithGallerySerializer(many=True)}
    )
    def get(self, request):
        produits = Produit.objects.filter(status='publié').order_by('-date_ajout')[:20]
        serializer = ProduitWithGallerySerializer(produits, many=True, context={'request': request})
        return Response(serializer.data)
