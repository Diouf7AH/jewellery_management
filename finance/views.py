# finance/views.py
from django.db import transaction
from django.db.models import Count, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from backend.roles import get_role_name
from backend.utils.helpers import resolve_bijouterie_for_user

from .models import Depense
from .serializers import DepenseSerializer

ROLE_ADMIN = "admin"
ROLE_MANAGER = "manager"


class DepenseCreateView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Créer une dépense",
        operation_description="""
        Crée une dépense directement payée.

        Règles :
        - Admin et Manager uniquement.
        - La dépense est automatiquement marquée comme PAYÉE.
        - Aucun ticket.
        - Aucun reçu.
        - Aucun impact sur le stock.
        - Utilisée uniquement pour l'audit et les rapports.
        """,
        request_body=DepenseSerializer,
        responses={201: DepenseSerializer},
        tags=["Finance - Dépenses"],
    )
    @transaction.atomic
    def post(self, request):
        user = request.user
        role = get_role_name(user)

        if role not in [ROLE_ADMIN, ROLE_MANAGER]:
            return Response({"detail": "Accès refusé."}, status=403)

        serializer = DepenseSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        bijouterie = resolve_bijouterie_for_user(user)

        depense = serializer.save(
            bijouterie=bijouterie,
            created_by=user,
            paid_by=user,
            paid_at=timezone.now(),
            status=Depense.STATUS_PAID,
        )

        return Response(DepenseSerializer(depense).data, status=201)


class DepenseListView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Liste des dépenses",
        operation_description="""
        Retourne la liste des dépenses.

        Règles :
        - Admin voit toutes les dépenses.
        - Manager voit seulement les dépenses de ses bijouteries.
        - Vendor et Cashier n'ont pas accès.
        """,
        manual_parameters=[
            openapi.Parameter("type_depense", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("status", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("start_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Format YYYY-MM-DD"),
            openapi.Parameter("end_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Format YYYY-MM-DD"),
            openapi.Parameter("bijouterie_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        ],
        responses={200: DepenseSerializer(many=True)},
        tags=["Finance - Dépenses"],
    )
    def get(self, request):
        user = request.user
        role = get_role_name(user)

        if role not in [ROLE_ADMIN, ROLE_MANAGER]:
            return Response({"detail": "Accès refusé."}, status=403)

        qs = Depense.objects.select_related(
            "bijouterie",
            "created_by",
            "paid_by",
            "cancelled_by",
        )

        if role == ROLE_MANAGER:
            qs = qs.filter(
                bijouterie__in=user.staff_manager_profile.bijouteries.all()
            )

        type_depense = request.query_params.get("type_depense")
        status_param = request.query_params.get("status")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        bijouterie_id = request.query_params.get("bijouterie_id")

        if type_depense:
            qs = qs.filter(type_depense=type_depense)

        if status_param:
            qs = qs.filter(status=status_param)

        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)

        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        if bijouterie_id:
            qs = qs.filter(bijouterie_id=bijouterie_id)

        serializer = DepenseSerializer(qs.order_by("-created_at"), many=True)
        return Response(serializer.data)


class CancelDepenseView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Annuler une dépense",
        operation_description="""
        Annule une dépense enregistrée.

        Règles :
        - Admin et Manager uniquement.
        - Une dépense annulée reste dans la base pour l'audit.
        - Aucun stock n'est touché.
        """,
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "reason": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Motif de l'annulation",
                )
            },
        ),
        responses={200: DepenseSerializer},
        tags=["Finance - Dépenses"],
    )
    @transaction.atomic
    def post(self, request, depense_id):
        user = request.user
        role = get_role_name(user)

        if role not in [ROLE_ADMIN, ROLE_MANAGER]:
            return Response({"detail": "Accès refusé."}, status=403)

        depense = get_object_or_404(
            Depense.objects.select_for_update(),
            id=depense_id,
        )

        if role == ROLE_MANAGER:
            manager_bijouteries = user.staff_manager_profile.bijouteries.all()
            if depense.bijouterie not in manager_bijouteries:
                return Response({"detail": "Accès refusé à cette dépense."}, status=403)

        if depense.status == Depense.STATUS_CANCELLED:
            return Response({"detail": "Cette dépense est déjà annulée."}, status=400)

        reason = request.data.get("reason", "")

        depense.status = Depense.STATUS_CANCELLED
        depense.cancelled_by = user
        depense.cancelled_at = timezone.now()
        depense.cancel_reason = reason
        depense.save()

        return Response(DepenseSerializer(depense).data)


class DepenseDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Dashboard des dépenses",
        operation_description="""
        Dashboard des dépenses.

        Règles :
        - Admin voit toutes les dépenses.
        - Manager voit seulement les dépenses de ses bijouteries.
        - Vendor et Cashier n'ont pas accès.
        """,
        manual_parameters=[
            openapi.Parameter("start_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Format YYYY-MM-DD"),
            openapi.Parameter("end_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Format YYYY-MM-DD"),
            openapi.Parameter("bijouterie_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        ],
        tags=["Finance - Dépenses"],
    )
    def get(self, request):
        user = request.user
        role = get_role_name(user)

        if role not in [ROLE_ADMIN, ROLE_MANAGER]:
            return Response({"detail": "Accès refusé."}, status=403)

        qs = Depense.objects.all()

        if role == ROLE_MANAGER:
            qs = qs.filter(
                bijouterie__in=user.staff_manager_profile.bijouteries.all()
            )

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        bijouterie_id = request.query_params.get("bijouterie_id")

        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)

        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        if bijouterie_id:
            qs = qs.filter(bijouterie_id=bijouterie_id)

        qs_paid = qs.filter(status=Depense.STATUS_PAID)
        qs_cancelled = qs.filter(status=Depense.STATUS_CANCELLED)

        total_depenses = qs_paid.aggregate(
            total=Coalesce(Sum("montant"), 0)
        )["total"]

        total_annule = qs_cancelled.aggregate(
            total=Coalesce(Sum("montant"), 0)
        )["total"]

        depenses_par_type = qs_paid.values("type_depense").annotate(
            total=Coalesce(Sum("montant"), 0),
            nombre=Count("id"),
        ).order_by("type_depense")

        depenses_par_bijouterie = qs_paid.values(
            "bijouterie_id",
            "bijouterie__nom",
        ).annotate(
            total=Coalesce(Sum("montant"), 0),
            nombre=Count("id"),
        ).order_by("bijouterie__nom")

        return Response({
            "total_depenses": total_depenses,
            "nombre_depenses": qs_paid.count(),
            "total_annule": total_annule,
            "nombre_annule": qs_cancelled.count(),
            "depenses_par_type": depenses_par_type,
            "depenses_par_bijouterie": depenses_par_bijouterie,
        })
        



class ExportDepensesExcelView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Exporter les dépenses en Excel",
        operation_description="""
        Exporte les dépenses en fichier Excel.

        Filtres disponibles :
        - start_date : date début, format YYYY-MM-DD
        - end_date : date fin, format YYYY-MM-DD
        - type_depense
        - status
        - bijouterie_id

        Règles :
        - Admin voit toutes les dépenses.
        - Manager voit seulement les dépenses de ses bijouteries.
        - Vendor et Cashier n'ont pas accès.
        """,
        manual_parameters=[
            openapi.Parameter("start_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="YYYY-MM-DD"),
            openapi.Parameter("end_date", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="YYYY-MM-DD"),
            openapi.Parameter("type_depense", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("status", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("bijouterie_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        ],
        tags=["Finance - Dépenses"],
    )
    def get(self, request):
        user = request.user
        role = get_role_name(user)

        if role not in [ROLE_ADMIN, ROLE_MANAGER]:
            return Response({"detail": "Accès refusé."}, status=403)

        qs = Depense.objects.select_related(
            "bijouterie",
            "created_by",
            "paid_by",
            "cancelled_by",
            "updated_by",
        )

        if role == ROLE_MANAGER:
            qs = qs.filter(
                bijouterie__in=user.staff_manager_profile.bijouteries.all()
            )

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        type_depense = request.query_params.get("type_depense")
        status_param = request.query_params.get("status")
        bijouterie_id = request.query_params.get("bijouterie_id")

        if start_date:
            parsed_start = parse_date(start_date)
            if not parsed_start:
                return Response({"detail": "start_date invalide. Format attendu YYYY-MM-DD."}, status=400)
            qs = qs.filter(created_at__date__gte=parsed_start)

        if end_date:
            parsed_end = parse_date(end_date)
            if not parsed_end:
                return Response({"detail": "end_date invalide. Format attendu YYYY-MM-DD."}, status=400)
            qs = qs.filter(created_at__date__lte=parsed_end)

        if type_depense:
            qs = qs.filter(type_depense=type_depense)

        if status_param:
            qs = qs.filter(status=status_param)

        if bijouterie_id:
            qs = qs.filter(bijouterie_id=bijouterie_id)

        qs = qs.order_by("-created_at")

        wb = Workbook()
        ws = wb.active
        ws.title = "Dépenses"

        headers = [
            "ID",
            "Date création",
            "Bijouterie",
            "Type",
            "Titre",
            "Description",
            "Montant",
            "Bénéficiaire",
            "Téléphone bénéficiaire",
            "Statut",
            "Créé par",
            "Payé par",
            "Date paiement",
            "Annulé par",
            "Date annulation",
            "Motif annulation",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        total_montant = 0

        for depense in qs:
            montant = depense.montant or 0

            if depense.status == Depense.STATUS_PAID:
                total_montant += montant

            ws.append([
                depense.id,
                depense.created_at.strftime("%d/%m/%Y %H:%M") if depense.created_at else "",
                depense.bijouterie.nom if depense.bijouterie else "",
                depense.get_type_depense_display(),
                depense.titre,
                depense.description or "",
                float(montant),
                depense.beneficiaire or "",
                depense.telephone_beneficiaire or "",
                depense.get_status_display(),
                str(depense.created_by) if depense.created_by else "",
                str(depense.paid_by) if depense.paid_by else "",
                depense.paid_at.strftime("%d/%m/%Y %H:%M") if depense.paid_at else "",
                str(depense.cancelled_by) if depense.cancelled_by else "",
                depense.cancelled_at.strftime("%d/%m/%Y %H:%M") if depense.cancelled_at else "",
                depense.cancel_reason or "",
            ])

        total_row = ws.max_row + 2
        ws.cell(row=total_row, column=6, value="TOTAL DÉPENSES PAYÉES").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=float(total_montant)).font = Font(bold=True)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                value = str(cell.value or "")
                max_length = max(max_length, len(value))

            ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "export_depenses.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response
    
    
