# inventory/views.py
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Optional, Set

from django.core.exceptions import ValidationError
from django.db.models import (Case, Count, DecimalField, ExpressionWrapper, F,
                              Prefetch, Q, Sum, Value, When)
from django.db.models.functions import Cast, Coalesce
from django.utils import timezone
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from rest_framework import status
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from backend.mixins import (GROUP_BY_CHOICES, ExportXlsxMixin,
                            aware_range_month, parse_month_or_default,
                            resolve_tz)
from backend.permissions import IsAdminOrManager
from backend.roles import ROLE_ADMIN, ROLE_MANAGER, get_role_name
from inventory.models import InventoryMovement, MovementType
from purchase.models import ProduitLine
from stock.models import VendorStock
from store.models import Bijouterie
from vendor.models import Vendor

from .serializers import (InventoryBijouterieSerializer,
                          InventoryVendorSerializer,
                          ProduitLineWithInventorySerializer)
from .utils import _b
from .utils import parse_date as _date
from .utils import parse_int as _int


class ProduitLineWithInventoryListView(ListAPIView):
    """
    Vue avancée FIFO / lot.

    Retourne les ProduitLine avec :
    - lot
    - achat
    - fournisseur
    - produit
    - stock agrégé
    - mouvements liés à la ProduitLine
    """

    permission_classes = [IsAuthenticated, IsAdminOrManager]
    serializer_class = ProduitLineWithInventorySerializer
    pagination_class = None

    @swagger_auto_schema(
        operation_id="listProduitLinesWithInventory",
        operation_summary="Lister les ProduitLine avec stock + mouvements",
        operation_description=(
            "Retourne une vue détaillée par ligne d'achat `ProduitLine`.\n\n"
            "Cette vue est utile pour :\n"
            "- suivre le FIFO\n"
            "- analyser les lots\n"
            "- relier achat → lot → produit → stock → mouvements\n\n"
            "Filtres disponibles :\n"
            "- `year`\n"
            "- `lot_id`\n"
            "- `produit_id`\n"
            "- `numero_lot`"
        ),
        manual_parameters=[
            openapi.Parameter(
                "year",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                description="Année du lot. Défaut : année courante.",
            ),
            openapi.Parameter(
                "lot_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                description="Filtrer par lot_id.",
            ),
            openapi.Parameter(
                "produit_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                description="Filtrer par produit_id.",
            ),
            openapi.Parameter(
                "numero_lot",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                description="Recherche partielle sur le numéro de lot.",
            ),
        ],
        tags=["Inventaire"],
        responses={200: ProduitLineWithInventorySerializer(many=True)},
    )
    def get_queryset(self):
        getf = self.request.query_params.get

        # -------------------------
        # Année
        # -------------------------
        year_raw = getf("year")

        if year_raw in (None, "", "null"):
            year = timezone.localdate().year
        else:
            try:
                year = int(year_raw)
            except ValueError:
                raise ValidationError({"year": "Doit être un entier."})

            if year < 2000 or year > 2100:
                raise ValidationError({"year": "Année invalide."})

        # -------------------------
        # Mouvements liés à ProduitLine
        # -------------------------
        movements_qs = (
            InventoryMovement.objects
            .select_related(
                "src_bijouterie",
                "dst_bijouterie",
                "vendor",
                "vendor__user",
                "facture",
                "vente",
                "created_by",
            )
            .order_by("-occurred_at", "-id")
        )

        qs = (
            ProduitLine.objects
            .select_related(
                "lot",
                "lot__achat",
                "lot__achat__fournisseur",
                "produit",
                "produit__categorie",
                "produit__marque",
                "produit__purete",
            )
            .prefetch_related(
                Prefetch(
                    "inventory_movements",
                    queryset=movements_qs,
                    to_attr="prefetched_movements",
                )
            )
            .annotate(
                quantite_allouee=Coalesce(
                    Sum("stocks__quantite_allouee"),
                    0,
                ),
                quantite_disponible_total=Coalesce(
                    Sum("stocks__quantite_disponible"),
                    0,
                ),
            )
            .filter(lot__received_at__year=year)
        )

        # -------------------------
        # Filtres
        # -------------------------
        lot_id = getf("lot_id")
        if lot_id:
            try:
                qs = qs.filter(lot_id=int(lot_id))
            except ValueError:
                raise ValidationError({"lot_id": "Doit être un entier."})

        produit_id = getf("produit_id")
        if produit_id:
            try:
                qs = qs.filter(produit_id=int(produit_id))
            except ValueError:
                raise ValidationError({"produit_id": "Doit être un entier."})

        numero_lot = (getf("numero_lot") or "").strip()
        if numero_lot:
            qs = qs.filter(lot__numero_lot__icontains=numero_lot)

        return qs.order_by("-lot__received_at", "lot__numero_lot", "id")


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

# ---------------------------------------------------------------------
# InventoryMovementListView
# ---------------------------------------------------------------------
def _int(v: Optional[str]) -> Optional[int]:
    if v in (None, "", "null"):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _date(v: Optional[str]) -> Optional[date]:
    if not v:
        return None
    try:
        return datetime.strptime(v, "%Y-%m-%d").date()
    except Exception:
        return None


def _signed_qty(movement_type, qty) -> int:
    q = int(qty or 0)
    if movement_type in {
        MovementType.SALE_OUT,
        MovementType.CANCEL_PURCHASE,
    }:
        return -q
    return q


class InventoryMovementListView(ExportXlsxMixin, APIView):
    """
    Journal détaillé des mouvements d’inventaire.

    - Admin   : voit tout
    - Manager : voit uniquement les mouvements de ses bijouteries
    - Export Excel : ?export=xlsx
    - Pagination simple : ?limit=200&offset=0
    """

    permission_classes = [IsAuthenticated]
    http_method_names = ["get"]

    @swagger_auto_schema(
        operation_id="inventoryMovementList",
        operation_summary="Journal des mouvements d’inventaire",
        operation_description=(
            "Retourne le journal détaillé des mouvements d’inventaire.\n\n"
            "### Rôles\n"
            "- Admin : tous les mouvements\n"
            "- Manager : mouvements de ses bijouteries uniquement\n\n"
            "### Filtres\n"
            "- q : recherche produit, SKU, lot, raison\n"
            "- date_from/date_to : YYYY-MM-DD\n"
            "- movement_types : CSV ex: PURCHASE_IN,SALE_OUT\n"
            "- produit_id, lot_id, achat_id, vendor_id, vente_id, facture_id\n"
            "- bijouterie_id, src_bijouterie_id, dst_bijouterie_id\n"
            "- src_bucket, dst_bucket\n"
            "- min_qty, max_qty\n\n"
            "### Export\n"
            "- export=xlsx"
        ),
        manual_parameters=[
            openapi.Parameter("q", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("date_from", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="YYYY-MM-DD"),
            openapi.Parameter("date_to", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="YYYY-MM-DD"),
            openapi.Parameter("movement_types", openapi.IN_QUERY, type=openapi.TYPE_STRING, description="CSV"),
            openapi.Parameter("produit_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("lot_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("lot_code", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("achat_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("vendor_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("vente_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("facture_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("src_bucket", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("dst_bucket", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("bijouterie_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("src_bijouterie_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("dst_bijouterie_id", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("min_qty", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("max_qty", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("ordering", openapi.IN_QUERY, type=openapi.TYPE_STRING),
            openapi.Parameter("export", openapi.IN_QUERY, type=openapi.TYPE_STRING, enum=["xlsx"]),
            openapi.Parameter("limit", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
            openapi.Parameter("offset", openapi.IN_QUERY, type=openapi.TYPE_INTEGER),
        ],
        tags=["Inventaire"],
    )
    def get(self, request):
        getf = request.GET.get
        role = get_role_name(request.user)

        if role not in {ROLE_ADMIN, ROLE_MANAGER}:
            return Response(
                {"detail": "Accès réservé aux admins et managers."},
                status=status.HTTP_403_FORBIDDEN,
            )

        qs = InventoryMovement.objects.select_related(
            "produit",
            "lot",
            "src_bijouterie",
            "dst_bijouterie",
            "created_by",
            "vendor",
            "vendor__user",
            "vente",
            "facture",
        )

        # -------------------------
        # Scope rôle
        # -------------------------
        if role == ROLE_ADMIN:
            bj_id = _int(getf("bijouterie_id"))
            if bj_id:
                qs = qs.filter(
                    Q(src_bijouterie_id=bj_id) |
                    Q(dst_bijouterie_id=bj_id)
                )

        else:
            manager_profile = getattr(request.user, "staff_manager_profile", None)

            if not manager_profile or (
                hasattr(manager_profile, "verifie") and not manager_profile.verifie
            ):
                return Response(
                    {"detail": "Profil manager invalide."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            bijouterie_ids = list(
                manager_profile.bijouteries.values_list("id", flat=True)
            )

            if not bijouterie_ids:
                return Response(
                    {"detail": "Ce manager n'a aucune bijouterie assignée."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            qs = qs.filter(
                Q(src_bijouterie_id__in=bijouterie_ids) |
                Q(dst_bijouterie_id__in=bijouterie_ids)
            )

        # -------------------------
        # Recherche texte
        # -------------------------
        q = (getf("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(produit__nom__icontains=q)
                | Q(produit__sku__icontains=q)
                | Q(lot__numero_lot__icontains=q)
                | Q(reason__icontains=q)
            )

        # -------------------------
        # Filtres IDs
        # -------------------------
        id_filters = {
            "produit_id": "produit_id",
            "lot_id": "lot_id",
            "achat_id": "achat_id",
            "vendor_id": "vendor_id",
            "vente_id": "vente_id",
            "facture_id": "facture_id",
            "src_bijouterie_id": "src_bijouterie_id",
            "dst_bijouterie_id": "dst_bijouterie_id",
        }

        for param, field in id_filters.items():
            value = _int(getf(param))
            if value:
                qs = qs.filter(**{field: value})

        lot_code = (getf("lot_code") or "").strip()
        if lot_code:
            qs = qs.filter(lot__numero_lot__iexact=lot_code)

        src_bucket = (getf("src_bucket") or "").strip()
        if src_bucket:
            qs = qs.filter(src_bucket=src_bucket)

        dst_bucket = (getf("dst_bucket") or "").strip()
        if dst_bucket:
            qs = qs.filter(dst_bucket=dst_bucket)

        # -------------------------
        # Types de mouvements
        # -------------------------
        types_csv = (getf("movement_types") or "").strip()
        if types_csv:
            types = [t.strip().upper() for t in types_csv.split(",") if t.strip()]
            allowed = {choice[0] for choice in MovementType.choices}
            bad = [t for t in types if t not in allowed]

            if bad:
                return Response(
                    {
                        "detail": (
                            f"movement_types invalide(s): {bad}. "
                            f"Allowed: {sorted(allowed)}"
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            qs = qs.filter(movement_type__in=types)

        # -------------------------
        # Dates
        # -------------------------
        date_from_raw = getf("date_from")
        date_to_raw = getf("date_to")

        date_from = _date(date_from_raw)
        date_to = _date(date_to_raw)

        if date_from_raw and not date_from:
            return Response(
                {"date_from": "Format invalide. Utiliser YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if date_to_raw and not date_to:
            return Response(
                {"date_to": "Format invalide. Utiliser YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if date_from and date_to and date_from > date_to:
            return Response(
                {"detail": "date_from doit être inférieur ou égal à date_to."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if date_from:
            qs = qs.filter(occurred_at__date__gte=date_from)

        if date_to:
            qs = qs.filter(occurred_at__date__lte=date_to)

        # -------------------------
        # Quantités
        # -------------------------
        min_qty = _int(getf("min_qty"))
        max_qty = _int(getf("max_qty"))

        if min_qty is not None:
            qs = qs.filter(qty__gte=min_qty)

        if max_qty is not None:
            qs = qs.filter(qty__lte=max_qty)

        # -------------------------
        # Tri
        # -------------------------
        ordering = (getf("ordering") or "-occurred_at").strip()

        allowed_ordering = {
            "occurred_at",
            "-occurred_at",
            "qty",
            "-qty",
            "movement_type",
            "-movement_type",
            "produit_nom",
            "-produit_nom",
            "id",
            "-id",
        }

        if ordering not in allowed_ordering:
            ordering = "-occurred_at"

        if "produit_nom" in ordering:
            ordering = ordering.replace("produit_nom", "produit__nom")

        qs = qs.order_by(ordering, "-id")

        values_fields = [
            "id",
            "occurred_at",
            "movement_type",
            "qty",
            "produit_id",
            "produit__nom",
            "produit__sku",
            "lot_id",
            "lot__numero_lot",
            "achat_id",
            "vendor_id",
            "vendor__user__email",
            "vente_id",
            "vente__numero_vente",
            "facture_id",
            "facture__numero_facture",
            "src_bucket",
            "src_bijouterie_id",
            "dst_bucket",
            "dst_bijouterie_id",
            "created_by_id",
            "created_by__email",
            "reason",
        ]

        # -------------------------
        # Export Excel
        # -------------------------
        if (getf("export") or "").lower() == "xlsx":
            rows = list(qs.values(*values_fields))

            bij_ids: Set[int] = set()
            for r in rows:
                if r.get("src_bijouterie_id"):
                    bij_ids.add(r["src_bijouterie_id"])
                if r.get("dst_bijouterie_id"):
                    bij_ids.add(r["dst_bijouterie_id"])

            bij_map = {
                b.id: b.nom
                for b in Bijouterie.objects.filter(id__in=bij_ids)
            } if bij_ids else {}

            wb = Workbook()
            ws = wb.active
            ws.title = "Mouvements"

            headers = [
                "id",
                "occurred_at",
                "movement_type",
                "produit_id",
                "produit_nom",
                "produit_sku",
                "lot_id",
                "lot_code",
                "qty",
                "signed_qty",
                "achat_id",
                "vendor_id",
                "vendor_email",
                "vente_id",
                "numero_vente",
                "facture_id",
                "numero_facture",
                "src_bucket",
                "src_bijouterie_id",
                "src_bijouterie_nom",
                "dst_bucket",
                "dst_bijouterie_id",
                "dst_bijouterie_nom",
                "created_by_id",
                "created_by_email",
                "reason",
            ]
            ws.append(headers)

            for r in rows:
                signed_qty = _signed_qty(r.get("movement_type"), r.get("qty"))

                ws.append([
                    r.get("id"),
                    r.get("occurred_at"),
                    r.get("movement_type"),
                    r.get("produit_id"),
                    r.get("produit__nom"),
                    r.get("produit__sku"),
                    r.get("lot_id"),
                    r.get("lot__numero_lot"),
                    r.get("qty"),
                    signed_qty,
                    r.get("achat_id"),
                    r.get("vendor_id"),
                    r.get("vendor__user__email"),
                    r.get("vente_id"),
                    r.get("vente__numero_vente"),
                    r.get("facture_id"),
                    r.get("facture__numero_facture"),
                    r.get("src_bucket"),
                    r.get("src_bijouterie_id"),
                    bij_map.get(r.get("src_bijouterie_id")),
                    r.get("dst_bucket"),
                    r.get("dst_bijouterie_id"),
                    bij_map.get(r.get("dst_bijouterie_id")),
                    r.get("created_by_id"),
                    r.get("created_by__email"),
                    r.get("reason"),
                ])

            self._autosize(ws)
            return self._xlsx_response(wb, "inventory_movements.xlsx")

        # -------------------------
        # Pagination JSON
        # -------------------------
        limit = _int(getf("limit")) or 200
        offset = _int(getf("offset")) or 0

        limit = max(1, min(limit, 2000))
        offset = max(0, offset)

        total = qs.count()
        page_qs = qs[offset: offset + limit]
        rows = list(page_qs.values(*values_fields))

        bij_ids: Set[int] = set()
        for r in rows:
            if r.get("src_bijouterie_id"):
                bij_ids.add(r["src_bijouterie_id"])
            if r.get("dst_bijouterie_id"):
                bij_ids.add(r["dst_bijouterie_id"])

        bij_map = {
            b.id: b.nom
            for b in Bijouterie.objects.filter(id__in=bij_ids)
        } if bij_ids else {}

        results = []

        for r in rows:
            results.append({
                "id": r.get("id"),
                "occurred_at": r.get("occurred_at"),
                "movement_type": r.get("movement_type"),

                "produit_id": r.get("produit_id"),
                "produit_nom": r.get("produit__nom"),
                "produit_sku": r.get("produit__sku"),

                "lot_id": r.get("lot_id"),
                "lot_code": r.get("lot__numero_lot"),

                "achat_id": r.get("achat_id"),

                "qty": r.get("qty"),
                "signed_qty": _signed_qty(r.get("movement_type"), r.get("qty")),

                "vendor_id": r.get("vendor_id"),
                "vendor_email": r.get("vendor__user__email"),

                "vente_id": r.get("vente_id"),
                "numero_vente": r.get("vente__numero_vente"),

                "facture_id": r.get("facture_id"),
                "numero_facture": r.get("facture__numero_facture"),

                "src_bucket": r.get("src_bucket"),
                "src_bijouterie_id": r.get("src_bijouterie_id"),
                "src_bijouterie_nom": bij_map.get(r.get("src_bijouterie_id")),

                "dst_bucket": r.get("dst_bucket"),
                "dst_bijouterie_id": r.get("dst_bijouterie_id"),
                "dst_bijouterie_nom": bij_map.get(r.get("dst_bijouterie_id")),

                "created_by_id": r.get("created_by_id"),
                "created_by_email": r.get("created_by__email"),

                "reason": r.get("reason"),
            })

        next_offset = offset + limit

        return Response(
            {
                "count": total,
                "limit": limit,
                "offset": offset,
                "next_offset": next_offset if next_offset < total else None,
                "results": results,
            },
            status=status.HTTP_200_OK,
        )




class InventoryBijouterieView(ExportXlsxMixin, APIView):
    """
    Résumé des entrées / sorties de stock par bijouterie.
    """
    permission_classes = [IsAuthenticated]
    http_method_names = ["get"]

    @swagger_auto_schema(
        operation_id="inventoryBijouterie",
        operation_summary="Résumé des entrées et sorties de stock par bijouterie",
        operation_description=(
            "Retourne un **résumé global des mouvements de stock par bijouterie**.\n\n"
            "Cette vue permet de connaître, pour chaque bijouterie :\n"
            "- les **entrées d'achat** (`purchase_in`)\n"
            "- les **annulations d'achat** (`cancel_purchase_out`)\n"
            "- les **affectations reçues depuis la réserve** (`allocate_in`)\n"
            "- les **transferts entrants** (`transfer_in`)\n"
            "- les **transferts sortants** (`transfer_out`)\n"
            "- les **sorties de vente** (`sale_out`)\n"
            "- les **retours clients** (`return_in`)\n"
            "- les **ajustements positifs** (`adjustment_in`)\n"
            "- les **ajustements négatifs** (`adjustment_out`)\n"
            "- le **solde net des mouvements** (`stock_net`)\n\n"
            "### Règles d'accès\n"
            "- **Admin** : peut voir toutes les bijouteries\n"
            "- **Manager** : ne voit que les bijouteries qui lui sont assignées\n\n"
            "### Filtres disponibles\n"
            "- `bijouterie_id` : limiter le résumé à une seule bijouterie\n"
            "- `produit_id` : limiter le résumé à un produit précis\n"
            "- `date_from` : date de début incluse (`YYYY-MM-DD`)\n"
            "- `date_to` : date de fin incluse (`YYYY-MM-DD`)\n"
            "- `export=xlsx` : exporter le résumé au format Excel\n\n"
            "### Notes métier\n"
            "- `stock_net` est un **solde de mouvements**, pas une photo instantanée du stock physique.\n"
            "- Pour voir l'état actuel du stock par lot/produit, utilise la vue d'**inventaire photo**.\n"
        ),
        manual_parameters=[
            openapi.Parameter(
                "bijouterie_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="ID d'une bijouterie pour limiter le résumé à cette bijouterie uniquement.",
            ),
            openapi.Parameter(
                "produit_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="ID d'un produit pour limiter le résumé à ce produit.",
            ),
            openapi.Parameter(
                "date_from",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
                description="Date de début incluse au format YYYY-MM-DD.",
            ),
            openapi.Parameter(
                "date_to",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
                description="Date de fin incluse au format YYYY-MM-DD.",
            ),
            openapi.Parameter(
                "export",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                enum=["xlsx"],
                description="Mettre `xlsx` pour exporter le résumé en fichier Excel.",
            ),
        ],
        responses={
            200: openapi.Response(
                description="Résumé des mouvements de stock par bijouterie.",
                schema=InventoryBijouterieSerializer(many=True),
                examples={
                    "application/json": [
                        {
                            "bijouterie_id": 1,
                            "bijouterie_nom": "Rio Gold Dakar",
                            "purchase_in": "120.00",
                            "cancel_purchase_out": "5.00",
                            "allocate_in": "20.00",
                            "transfer_in": "8.00",
                            "transfer_out": "3.00",
                            "sale_out": "42.00",
                            "return_in": "2.00",
                            "adjustment_in": "1.00",
                            "adjustment_out": "0.00",
                            "stock_net": "101.00",
                        }
                    ]
                },
            ),
            400: openapi.Response(
                description="Paramètres invalides.",
                examples={
                    "application/json": {
                        "date_from": "Format invalide. Utiliser YYYY-MM-DD."
                    }
                },
            ),
            403: openapi.Response(
                description="Accès refusé.",
                examples={
                    "application/json": {
                        "detail": "Accès refusé."
                    }
                },
            ),
        },
        tags=["Inventaire / Résumés"],
    )
    def get(self, request):
        role = get_role_name(request.user)
        if role not in {ROLE_ADMIN, ROLE_MANAGER}:
            return Response({"detail": "Accès refusé."}, status=status.HTTP_403_FORBIDDEN)

        getf = request.GET.get
        bijouterie_id = _int(getf("bijouterie_id"))
        produit_id = _int(getf("produit_id"))
        date_from = _date(getf("date_from"))
        date_to = _date(getf("date_to"))

        if getf("date_from") and not date_from:
            return Response(
                {"date_from": "Format invalide. Utiliser YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if getf("date_to") and not date_to:
            return Response(
                {"date_to": "Format invalide. Utiliser YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_from and date_to and date_from > date_to:
            return Response(
                {"detail": "date_from doit être <= date_to."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = InventoryMovement.objects.select_related(
            "src_bijouterie",
            "dst_bijouterie",
            "produit",
        )

        allowed_ids = None
        if role == ROLE_MANAGER:
            mp = getattr(request.user, "staff_manager_profile", None)
            if not mp or (hasattr(mp, "verifie") and not mp.verifie):
                return Response(
                    {"detail": "Profil manager invalide."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            allowed_ids = list(mp.bijouteries.values_list("id", flat=True))
            if not allowed_ids:
                return Response(
                    {"detail": "Ce manager n'a aucune bijouterie assignée."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            qs = qs.filter(
                Q(src_bijouterie_id__in=allowed_ids) |
                Q(dst_bijouterie_id__in=allowed_ids)
            )

        if bijouterie_id:
            qs = qs.filter(
                Q(src_bijouterie_id=bijouterie_id) |
                Q(dst_bijouterie_id=bijouterie_id)
            )

        if produit_id:
            qs = qs.filter(produit_id=produit_id)

        if date_from:
            qs = qs.filter(occurred_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(occurred_at__date__lte=date_to)

        bijouteries = Bijouterie.objects.all()
        if role == ROLE_MANAGER and allowed_ids is not None:
            bijouteries = bijouteries.filter(id__in=allowed_ids)
        if bijouterie_id:
            bijouteries = bijouteries.filter(id=bijouterie_id)

        rows = []
        for bj in bijouteries.order_by("nom"):
            in_qs = qs.filter(dst_bijouterie_id=bj.id)
            out_qs = qs.filter(src_bijouterie_id=bj.id)

            purchase_in = in_qs.filter(
                movement_type=MovementType.PURCHASE_IN
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            cancel_purchase_out = out_qs.filter(
                movement_type=MovementType.CANCEL_PURCHASE
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            allocate_in = in_qs.filter(
                movement_type=MovementType.ALLOCATE
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            transfer_in = in_qs.filter(
                movement_type=MovementType.TRANSFER
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            transfer_out = out_qs.filter(
                movement_type=MovementType.TRANSFER
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            sale_out = out_qs.filter(
                movement_type=MovementType.SALE_OUT
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            return_in = in_qs.filter(
                movement_type=MovementType.RETURN_IN
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            adjustment_in = in_qs.filter(
                movement_type=MovementType.ADJUSTMENT
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            adjustment_out = out_qs.filter(
                movement_type=MovementType.ADJUSTMENT
            ).aggregate(
                v=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["v"]

            stock_net = (
                purchase_in
                + allocate_in
                + transfer_in
                + return_in
                + adjustment_in
                - cancel_purchase_out
                - transfer_out
                - sale_out
                - adjustment_out
            )

            rows.append({
                "bijouterie_id": bj.id,
                "bijouterie_nom": bj.nom,
                "purchase_in": purchase_in,
                "cancel_purchase_out": cancel_purchase_out,
                "allocate_in": allocate_in,
                "transfer_in": transfer_in,
                "transfer_out": transfer_out,
                "sale_out": sale_out,
                "return_in": return_in,
                "adjustment_in": adjustment_in,
                "adjustment_out": adjustment_out,
                "stock_net": stock_net,
            })

        if (getf("export") or "").lower() == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Résumé bijouteries"

            headers = [
                "bijouterie_id",
                "bijouterie_nom",
                "purchase_in",
                "cancel_purchase_out",
                "allocate_in",
                "transfer_in",
                "transfer_out",
                "sale_out",
                "return_in",
                "adjustment_in",
                "adjustment_out",
                "stock_net",
            ]
            ws.append(headers)

            for row in rows:
                ws.append([row.get(h) for h in headers])

            self._autosize(ws)
            return self._xlsx_response(wb, "inventory_bijouterie.xlsx")

        serializer = InventoryBijouterieSerializer(rows, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class InventoryVendorView(ExportXlsxMixin, APIView):
    """
    Résumé du stock et des mouvements par vendeur.
    """
    permission_classes = [IsAuthenticated]
    http_method_names = ["get"]

    @swagger_auto_schema(
        operation_id="inventoryVendor",
        operation_summary="Résumé du stock par vendeur",
        operation_description=(
            "Retourne un **résumé du stock et des mouvements par vendeur**.\n\n"
            "Cette vue permet de connaître, pour chaque vendeur :\n"
            "- le **stock affecté reçu** (`vendor_assign_in`)\n"
            "- les **sorties de vente du vendeur** (`sale_out_vendor`)\n"
            "- les **retours/restaurations éventuels** (`return_in_vendor`)\n"
            "- le **stock restant vendeur** (`stock_restant`)\n\n"
            "### Règles d'accès\n"
            "- **Admin** : peut voir tous les vendeurs\n"
            "- **Manager** : ne voit que les vendeurs de ses bijouteries\n\n"
            "### Filtres disponibles\n"
            "- `vendor_id` : limiter à un vendeur précis\n"
            "- `bijouterie_id` : limiter aux vendeurs d'une bijouterie\n"
            "- `produit_id` : limiter à un produit précis\n"
            "- `date_from` : date de début incluse (`YYYY-MM-DD`)\n"
            "- `date_to` : date de fin incluse (`YYYY-MM-DD`)\n"
            "- `export=xlsx` : exporter le résumé en Excel\n\n"
            "### Notes métier\n"
            "- `vendor_assign_in` correspond aux mouvements `VENDOR_ASSIGN` liés au vendeur.\n"
            "- `sale_out_vendor` correspond aux mouvements `SALE_OUT` liés au vendeur.\n"
            "- `stock_restant` est calculé depuis `VendorStock` : `quantite_allouee - quantite_vendue`.\n"
        ),
        manual_parameters=[
            openapi.Parameter(
                "vendor_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="ID d'un vendeur pour limiter le résumé à ce vendeur uniquement.",
            ),
            openapi.Parameter(
                "bijouterie_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="ID d'une bijouterie pour limiter aux vendeurs de cette bijouterie.",
            ),
            openapi.Parameter(
                "produit_id",
                openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="ID d'un produit pour limiter le résumé à ce produit.",
            ),
            openapi.Parameter(
                "date_from",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
                description="Date de début incluse au format YYYY-MM-DD.",
            ),
            openapi.Parameter(
                "date_to",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
                description="Date de fin incluse au format YYYY-MM-DD.",
            ),
            openapi.Parameter(
                "export",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
                enum=["xlsx"],
                description="Mettre `xlsx` pour exporter le résumé vendeur en fichier Excel.",
            ),
        ],
        responses={
            200: openapi.Response(
                description="Résumé du stock par vendeur.",
                schema=InventoryVendorSerializer(many=True),
                examples={
                    "application/json": [
                        {
                            "vendor_id": 3,
                            "vendor_nom": "Amadou Diop",
                            "vendor_email": "amadou@riogold.sn",
                            "bijouterie_id": 1,
                            "bijouterie_nom": "Rio Gold Dakar",
                            "vendor_assign_in": "80.00",
                            "sale_out_vendor": "25.00",
                            "return_in_vendor": "2.00",
                            "stock_restant": "57.00",
                        }
                    ]
                },
            ),
            400: openapi.Response(
                description="Paramètres invalides.",
                examples={
                    "application/json": {
                        "vendor_id": "Doit être un entier."
                    }
                },
            ),
            403: openapi.Response(
                description="Accès refusé.",
                examples={
                    "application/json": {
                        "detail": "Accès refusé."
                    }
                },
            ),
        },
        tags=["Inventaire / Résumés"],
    )
    def get(self, request):
        role = get_role_name(request.user)
        if role not in {ROLE_ADMIN, ROLE_MANAGER}:
            return Response({"detail": "Accès refusé."}, status=status.HTTP_403_FORBIDDEN)

        getf = request.GET.get
        vendor_id = _int(getf("vendor_id"))
        bijouterie_id = _int(getf("bijouterie_id"))
        produit_id = _int(getf("produit_id"))
        date_from = _date(getf("date_from"))
        date_to = _date(getf("date_to"))

        if getf("date_from") and not date_from:
            return Response(
                {"date_from": "Format invalide. Utiliser YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if getf("date_to") and not date_to:
            return Response(
                {"date_to": "Format invalide. Utiliser YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_from and date_to and date_from > date_to:
            return Response(
                {"detail": "date_from doit être <= date_to."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        vendors = Vendor.objects.select_related("user", "bijouterie").all()

        allowed_ids = None
        if role == ROLE_MANAGER:
            mp = getattr(request.user, "staff_manager_profile", None)
            if not mp or (hasattr(mp, "verifie") and not mp.verifie):
                return Response(
                    {"detail": "Profil manager invalide."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            allowed_ids = list(mp.bijouteries.values_list("id", flat=True))
            if not allowed_ids:
                return Response(
                    {"detail": "Ce manager n'a aucune bijouterie assignée."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            vendors = vendors.filter(bijouterie_id__in=allowed_ids)

        if vendor_id:
            vendors = vendors.filter(id=vendor_id)

        if bijouterie_id:
            vendors = vendors.filter(bijouterie_id=bijouterie_id)

        move_qs = InventoryMovement.objects.select_related(
            "vendor",
            "vendor__user",
            "vendor__bijouterie",
            "produit",
        )
        stock_qs = VendorStock.objects.select_related(
            "vendor",
            "vendor__user",
            "vendor__bijouterie",
            "produit_line",
            "produit_line__produit",
        )

        if role == ROLE_MANAGER and allowed_ids is not None:
            move_qs = move_qs.filter(vendor__bijouterie_id__in=allowed_ids)
            stock_qs = stock_qs.filter(vendor__bijouterie_id__in=allowed_ids)

        if vendor_id:
            move_qs = move_qs.filter(vendor_id=vendor_id)
            stock_qs = stock_qs.filter(vendor_id=vendor_id)

        if bijouterie_id:
            move_qs = move_qs.filter(vendor__bijouterie_id=bijouterie_id)
            stock_qs = stock_qs.filter(vendor__bijouterie_id=bijouterie_id)

        if produit_id:
            move_qs = move_qs.filter(produit_id=produit_id)
            stock_qs = stock_qs.filter(produit_line__produit_id=produit_id)

        if date_from:
            move_qs = move_qs.filter(occurred_at__date__gte=date_from)
        if date_to:
            move_qs = move_qs.filter(occurred_at__date__lte=date_to)

        rows = []
        for v in vendors.order_by("id"):
            vendor_assign_in = move_qs.filter(
                vendor_id=v.id,
                movement_type=MovementType.VENDOR_ASSIGN,
            ).aggregate(
                s=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["s"]

            sale_out_vendor = move_qs.filter(
                vendor_id=v.id,
                movement_type=MovementType.SALE_OUT,
            ).aggregate(
                s=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["s"]

            return_in_vendor = move_qs.filter(
                vendor_id=v.id,
                movement_type=MovementType.RETURN_IN,
            ).aggregate(
                s=Coalesce(
                    Sum("qty"),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["s"]

            stock_restant = stock_qs.filter(
                vendor_id=v.id
            ).aggregate(
                s=Coalesce(
                    Sum(
                        ExpressionWrapper(
                            F("quantite_allouee") - F("quantite_vendue"),
                            output_field=DecimalField(max_digits=18, decimal_places=2),
                        )
                    ),
                    Value(0),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )["s"]

            full_name = ""
            if getattr(v, "user", None):
                full_name = f"{(v.user.first_name or '').strip()} {(v.user.last_name or '').strip()}".strip()

            rows.append({
                "vendor_id": v.id,
                "vendor_nom": full_name or getattr(v, "nom", "") or "",
                "vendor_email": getattr(getattr(v, "user", None), "email", None),
                "bijouterie_id": getattr(v, "bijouterie_id", None),
                "bijouterie_nom": getattr(getattr(v, "bijouterie", None), "nom", None),
                "vendor_assign_in": vendor_assign_in,
                "sale_out_vendor": sale_out_vendor,
                "return_in_vendor": return_in_vendor,
                "stock_restant": stock_restant,
            })

        if (getf("export") or "").lower() == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Résumé vendeurs"

            headers = [
                "vendor_id",
                "vendor_nom",
                "vendor_email",
                "bijouterie_id",
                "bijouterie_nom",
                "vendor_assign_in",
                "sale_out_vendor",
                "return_in_vendor",
                "stock_restant",
            ]
            ws.append(headers)

            for row in rows:
                ws.append([row.get(h) for h in headers])

            self._autosize(ws)
            return self._xlsx_response(wb, "inventory_vendor.xlsx")

        serializer = InventoryVendorSerializer(rows, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    





