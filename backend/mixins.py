# backend/mixins.py

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============================================================
# Helpers date / timezone
# ============================================================

def aware_range_month(
    year: int,
    month: int,
    tz,
):
    """
    Retourne une période mensuelle sous la forme [start, end).

    Exemple :
        2026-01
        start = 2026-01-01 00:00
        end   = 2026-02-01 00:00

    La borne de fin est exclusive.
    """

    try:
        year = int(year)
        month = int(month)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "L'année et le mois doivent être des entiers."
        ) from exc

    if month < 1 or month > 12:
        raise ValueError(
            "Le mois doit être compris entre 1 et 12."
        )

    start_date = date(
        year,
        month,
        1,
    )

    if month == 12:
        end_date = date(
            year + 1,
            1,
            1,
        )
    else:
        end_date = date(
            year,
            month + 1,
            1,
        )

    start_dt = timezone.make_aware(
        datetime.combine(
            start_date,
            datetime.min.time(),
        ),
        timezone=tz,
    )

    end_dt = timezone.make_aware(
        datetime.combine(
            end_date,
            datetime.min.time(),
        ),
        timezone=tz,
    )

    return start_dt, end_dt


def parse_month_or_default(
    mois_str: str | None,
):
    """
    Parse le paramètre mois=YYYY-MM.

    Si le paramètre est absent, utilise le mois courant.

    Retourne :
        (
            annee,
            mois_num,
            mois_str_normalise,
        )
    """

    today = timezone.localdate()

    if not mois_str:
        return (
            today.year,
            today.month,
            today.strftime("%Y-%m"),
        )

    normalized = str(mois_str).strip()

    try:
        parsed = datetime.strptime(
            normalized,
            "%Y-%m",
        )
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "Format invalide. Utiliser mois=YYYY-MM."
        ) from exc

    normalized = parsed.strftime("%Y-%m")

    return (
        parsed.year,
        parsed.month,
        normalized,
    )


def resolve_tz(
    tz_name: str | None,
):
    """
    Résout une timezone IANA.

    Exemple :
        Africa/Dakar

    Si aucune timezone n'est fournie,
    retourne la timezone active du projet.
    """

    if not tz_name:
        return timezone.get_current_timezone()

    normalized = str(tz_name).strip()

    if not normalized:
        return timezone.get_current_timezone()

    try:
        return ZoneInfo(normalized)
    except ZoneInfoNotFoundError as exc:
        raise ValueError(
            "Timezone invalide. Exemple : tz=Africa/Dakar."
        ) from exc


# ============================================================
# Constantes de rapport
# ============================================================

GROUP_BY_CHOICES = {
    "lines",
    "day",
    "produit",
    "vendor",
    "bijouterie",
}


ORDERING_MAP = {
    # Jour
    "date": "date",
    "-date": "-date",
    "total_ht": "total_ht",
    "-total_ht": "-total_ht",
    "total_ttc": "total_ttc",
    "-total_ttc": "-total_ttc",
    "quantite": "quantite",
    "-quantite": "-quantite",

    # Produit
    "produit": "produit",
    "-produit": "-produit",

    # Vendeur
    "vendor_email": "vendor_email",
    "-vendor_email": "-vendor_email",

    # Bijouterie
    "bijouterie_nom": "bijouterie_nom",
    "-bijouterie_nom": "-bijouterie_nom",
}


# ============================================================
# Export Excel
# ============================================================

class ExportXlsxMixin:
    """
    Mixin permettant de renvoyer un classeur Excel.

    Exemple :

        class MaVue(ExportXlsxMixin, APIView):
            def get(self, request):
                wb = Workbook()
                return self._xlsx_response(
                    wb,
                    "rapport.xlsx",
                )
    """

    def _xlsx_response(
        self,
        wb: Workbook,
        filename: str,
    ) -> HttpResponse:
        """
        Génère une réponse HTTP contenant un fichier XLSX.
        """

        safe_filename = (
            str(filename or "export.xlsx")
            .replace('"', "")
            .replace("\n", "")
            .replace("\r", "")
            .strip()
        )

        if not safe_filename.lower().endswith(".xlsx"):
            safe_filename = f"{safe_filename}.xlsx"

        output = BytesIO()

        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{safe_filename}"'
        )

        return response

    def _autosize(
        self,
        ws,
        *,
        min_width: int = 10,
        max_width: int = 50,
    ):
        """
        Ajuste automatiquement la largeur des colonnes.
        """

        if min_width <= 0:
            raise ValueError(
                "min_width doit être supérieur à zéro."
            )

        if max_width < min_width:
            raise ValueError(
                "max_width doit être supérieur ou égal à min_width."
            )

        for column_cells in ws.columns:
            first_cell = column_cells[0]

            column_letter = get_column_letter(
                first_cell.column
            )

            max_length = 0

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(value)),
                )

            adjusted_width = max(
                min_width,
                min(
                    max_length + 2,
                    max_width,
                ),
            )

            ws.column_dimensions[
                column_letter
            ].width = adjusted_width
    
    