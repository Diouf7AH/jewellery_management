# compte_depot/views.py

import uuid
from decimal import Decimal

from django.conf import settings
from django.db import transaction
from django.db.models import Count, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from backend.bijouteries import get_user_bijouteries
from backend.renderers import UserRenderer
from backend.roles import (ROLE_ADMIN, ROLE_CASHIER, ROLE_MANAGER, ROLE_VENDOR,
                           get_role_name)
from backend.utils.helpers import (resolve_bijouterie_for_user,
                                   user_can_access_bijouterie)
from compte_depot.models import CompteDepot, CompteDepotTransaction
from compte_depot.notifications import (send_compte_created_notification,
                                        send_compte_depot_notification)

from .models import ClientDepot, CompteDepot, CompteDepotTransaction
from .pdf import generate_transaction_ticket_80mm_pdf
from .serializers import (ClientDepotSerializer, CompteDepotSerializer,
                          CompteDepotTelephoneTransactionSerializer,
                          CompteDepotTransactionSerializer,
                          CreateOrDepotCompteSerializer)
from .services import effectuer_depot, effectuer_retrait


class CreateOrDepotCompteView(APIView):
    permission_classes = [IsAuthenticated]

    ALLOWED_ROLES = {
        ROLE_ADMIN,
        ROLE_MANAGER,
        ROLE_CASHIER,
    }

    @swagger_auto_schema(
        operation_id="createOrDepotCompte",
        operation_summary="Créer ou alimenter un compte dépôt",
        operation_description=(
            "Crée un client et un compte dépôt si nécessaire.\n\n"
            "Si le client possède déjà un compte via son téléphone, "
            "effectue directement un dépôt sur ce compte.\n\n"
            "### Règles d'accès\n"
            "- Admin : autorisé, `bijouterie_id` obligatoire.\n"
            "- Manager : autorisé, `bijouterie_id` obligatoire et doit appartenir à ses bijouteries.\n"
            "- Caissier : autorisé, sa bijouterie est utilisée automatiquement.\n"
            "- Vendor : non autorisé.\n"
            "- Buyer : non autorisé.\n\n"
            "### Comportement\n"
            "- Client inexistant : création client + compte + dépôt initial.\n"
            "- Client existant sans compte : création du compte + dépôt initial.\n"
            "- Client avec compte existant : dépôt directement sur le compte.\n"
        ),
        request_body=CreateOrDepotCompteSerializer,
        responses={
            200: openapi.Response(
                description="Compte existant détecté et dépôt effectué.",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(
                            type=openapi.TYPE_STRING,
                        ),
                        "operation_type": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            example="deposit_existing_account",
                        ),
                        "client": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                        "bijouterie": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                        "compte": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                        "transaction": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                    },
                ),
            ),

            201: openapi.Response(
                description="Nouveau compte créé puis crédité avec succès.",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(
                            type=openapi.TYPE_STRING,
                        ),
                        "operation_type": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            example="create_client_account_and_deposit",
                        ),
                        "client": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                        "bijouterie": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                        "compte": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                        "transaction": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                        ),
                    },
                ),
            ),

            400: openapi.Response(
                description=(
                    "Données invalides, montant incorrect "
                    "ou bijouterie_id manquant."
                )
            ),

            403: openapi.Response(
                description=(
                    "Accès refusé ou bijouterie non accessible."
                )
            ),
        },
        tags=["compte dépôt"],
    )
    @transaction.atomic
    def post(self, request):
        role = get_role_name(request.user)

        if role not in self.ALLOWED_ROLES:
            return Response(
                {"detail": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = CreateOrDepotCompteSerializer(
            data=request.data
        )
        serializer.is_valid(raise_exception=True)

        client_data = serializer.validated_data["client"]
        montant = serializer.validated_data["montant"]
        telephone = client_data["telephone"]
        bijouterie_id = serializer.validated_data.get(
            "bijouterie_id"
        )

        # =====================================================
        # BIJOUTERIES ACCESSIBLES
        # =====================================================

        bijouteries = get_user_bijouteries(
            request.user
        )

        if not bijouteries.exists():
            return Response(
                {
                    "detail":
                    "Aucune bijouterie accessible pour cet utilisateur."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # =====================================================
        # CASHIER
        # =====================================================

        if role == ROLE_CASHIER:
            user_bijouterie = bijouteries.first()

            if not user_bijouterie:
                return Response(
                    {
                        "detail":
                        "Aucune bijouterie associée au caissier."
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

        # =====================================================
        # ADMIN / MANAGER
        # =====================================================

        else:
            if not bijouterie_id:
                return Response(
                    {
                        "detail":
                        "Le champ bijouterie_id est obligatoire "
                        "pour l'admin et le manager."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            user_bijouterie = (
                bijouteries
                .filter(id=bijouterie_id)
                .first()
            )

            if not user_bijouterie:
                return Response(
                    {
                        "detail":
                        "Cette bijouterie est introuvable "
                        "ou vous n'y avez pas accès."
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

        # =====================================================
        # RECHERCHE CLIENT
        # =====================================================

        client = (
            ClientDepot.objects
            .select_for_update()
            .filter(telephone=telephone)
            .first()
        )

        # =====================================================
        # CAS 1 : CLIENT EXISTANT
        # =====================================================

        if client:
            if not client.bijouterie_id:
                return Response(
                    {
                        "detail":
                        "Ce client n'est lié à aucune bijouterie."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if client.bijouterie_id != user_bijouterie.id:
                return Response(
                    {
                        "detail":
                        "Ce client appartient à une autre bijouterie."
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            compte = (
                CompteDepot.objects
                .select_for_update()
                .filter(client=client)
                .first()
            )

            # =================================================
            # CAS 1A : COMPTE EXISTANT
            # =================================================

            if compte:
                tx = effectuer_depot(
                    compte_id=compte.id,
                    montant=montant,
                    user=request.user,
                    reference="DEPOT_COMPTE_EXISTANT",
                    commentaire=(
                        "Dépôt effectué sur compte existant "
                        "via téléphone client."
                    ),
                )

                transaction.on_commit(
                    lambda tx=tx:
                    send_compte_depot_notification(tx)
                )

                compte.refresh_from_db()

                return Response(
                    {
                        "message":
                        "Compte existant détecté, dépôt effectué avec succès.",

                        "operation_type":
                        "deposit_existing_account",

                        "client":
                        ClientDepotSerializer(client).data,

                        "bijouterie": {
                            "id": user_bijouterie.id,
                            "nom": user_bijouterie.nom,
                        },

                        "compte": {
                            "id": compte.id,
                            "numero_compte": compte.numero_compte,
                            "solde": str(compte.solde),
                            "created_at": compte.created_at,
                        },

                        "transaction":
                        CompteDepotTransactionSerializer(tx).data,
                    },
                    status=status.HTTP_200_OK,
                )

            # =================================================
            # CAS 1B : CLIENT EXISTANT SANS COMPTE
            # =================================================

            numero = self.generer_numero_compte(
                telephone
            )

            compte = CompteDepot.objects.create(
                client=client,
                numero_compte=numero,
                created_by=request.user,
                solde=Decimal("0.00"),
            )

            tx = effectuer_depot(
                compte_id=compte.id,
                montant=montant,
                user=request.user,
                reference="OUVERTURE_COMPTE_CLIENT_EXISTANT",
                commentaire=(
                    "Nouveau compte créé puis crédité "
                    "pour un client existant."
                ),
            )

            transaction.on_commit(
                lambda compte=compte, montant=montant:
                send_compte_created_notification(
                    compte,
                    montant,
                )
            )

            compte.refresh_from_db()

            return Response(
                {
                    "message":
                    "Nouveau compte créé puis crédité avec succès.",

                    "operation_type":
                    "create_account_and_deposit",

                    "client":
                    ClientDepotSerializer(client).data,

                    "bijouterie": {
                        "id": user_bijouterie.id,
                        "nom": user_bijouterie.nom,
                    },

                    "compte": {
                        "id": compte.id,
                        "numero_compte": compte.numero_compte,
                        "solde": str(compte.solde),
                        "created_at": compte.created_at,
                    },

                    "transaction":
                    CompteDepotTransactionSerializer(tx).data,
                },
                status=status.HTTP_201_CREATED,
            )

        # =====================================================
        # CAS 2 : NOUVEAU CLIENT
        # =====================================================

        client = ClientDepot.objects.create(
            **client_data,
            bijouterie=user_bijouterie,
        )

        numero = self.generer_numero_compte(
            telephone
        )

        compte = CompteDepot.objects.create(
            client=client,
            numero_compte=numero,
            created_by=request.user,
            solde=Decimal("0.00"),
        )

        tx = effectuer_depot(
            compte_id=compte.id,
            montant=montant,
            user=request.user,
            reference="OUVERTURE_NOUVEAU_COMPTE",
            commentaire=(
                "Nouveau client créé, "
                "compte ouvert et crédité."
            ),
        )

        transaction.on_commit(
            lambda compte=compte, montant=montant:
            send_compte_created_notification(
                compte,
                montant,
            )
        )

        compte.refresh_from_db()

        return Response(
            {
                "message":
                "Nouveau client créé, compte ouvert et crédité avec succès.",

                "operation_type":
                "create_client_account_and_deposit",

                "client":
                ClientDepotSerializer(client).data,

                "bijouterie": {
                    "id": user_bijouterie.id,
                    "nom": user_bijouterie.nom,
                },

                "compte": {
                    "id": compte.id,
                    "numero_compte": compte.numero_compte,
                    "solde": str(compte.solde),
                    "created_at": compte.created_at,
                },

                "transaction":
                CompteDepotTransactionSerializer(tx).data,
            },
            status=status.HTTP_201_CREATED,
        )

    def generer_numero_compte(self, telephone):
        date_str = timezone.localtime(
            timezone.now()
        ).strftime("%y%m")

        prefix = f"{telephone}-{date_str}"[:25]

        if not CompteDepot.objects.filter(
            numero_compte=prefix
        ).exists():
            return prefix

        for _ in range(10):
            suffix = uuid.uuid4().hex[:4].upper()

            numero = f"{prefix}-{suffix}"

            if not CompteDepot.objects.filter(
                numero_compte=numero
            ).exists():
                return numero

        raise RuntimeError(
            "Impossible de générer un numéro de compte unique."
        )


class DepotView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_id="depotCompteDepot",
        operation_summary="Effectuer un dépôt sur un compte dépôt",
        operation_description=(
            "Effectue un dépôt sur un compte dépôt via le téléphone du client.\n\n"
            "### Règles\n"
            "- Admin : toutes les bijouteries.\n"
            "- Manager : uniquement ses bijouteries.\n"
            "- Caissier : uniquement sa bijouterie.\n"
            "- Vendor : non autorisé.\n"
            "- Buyer : non autorisé."
        ),
        request_body=CompteDepotTelephoneTransactionSerializer,
        responses={
            201: openapi.Response(
                description="Dépôt effectué avec succès",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            example="Dépôt effectué avec succès.",
                        ),
                        "client": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "nom": openapi.Schema(
                                    type=openapi.TYPE_STRING
                                ),
                                "prenom": openapi.Schema(
                                    type=openapi.TYPE_STRING
                                ),
                                "telephone": openapi.Schema(
                                    type=openapi.TYPE_STRING
                                ),
                            },
                        ),
                        "bijouterie": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "id": openapi.Schema(
                                    type=openapi.TYPE_INTEGER
                                ),
                                "nom": openapi.Schema(
                                    type=openapi.TYPE_STRING
                                ),
                            },
                        ),
                        "transaction": openapi.Schema(
                            type=openapi.TYPE_OBJECT
                        ),
                        "compte": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "numero_compte": openapi.Schema(
                                    type=openapi.TYPE_STRING
                                ),
                                "nouveau_solde": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    example="60000.00",
                                ),
                            },
                        ),
                        "receipt_url": openapi.Schema(
                            type=openapi.TYPE_STRING
                        ),
                    },
                ),
            ),
            400: openapi.Response(
                description="Données invalides ou compte sans bijouterie"
            ),
            403: openapi.Response(
                description="Accès refusé ou bijouterie non autorisée"
            ),
            404: openapi.Response(
                description="Compte introuvable"
            ),
        },
        tags=["compte dépôt"],
    )
    @transaction.atomic
    def post(self, request):
        
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {
                    "message": (
                        "Seul l'admin, le manager ou le caissier "
                        "peut effectuer un dépôt."
                    )
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = CompteDepotTelephoneTransactionSerializer(
            data=request.data
        )
        serializer.is_valid(raise_exception=True)

        telephone = serializer.validated_data["telephone"]

        try:
            compte = (
                CompteDepot.objects
                .select_for_update()
                .select_related(
                    "client",
                    "client__bijouterie",
                )
                .get(client__telephone=telephone)
            )
        except CompteDepot.DoesNotExist:
            return Response(
                {"detail": "Compte introuvable."},
                status=status.HTTP_404_NOT_FOUND,
            )

        client_bijouterie = getattr(
            compte.client,
            "bijouterie",
            None,
        )

        if not client_bijouterie:
            return Response(
                {
                    "message":
                    "Ce compte dépôt n'est lié à aucune bijouterie."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not user_can_access_bijouterie(
            request.user,
            client_bijouterie,
        ):
            return Response(
                {
                    "message":
                    "Vous ne pouvez pas effectuer un dépôt "
                    "sur ce compte."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        tx = effectuer_depot(
            compte_id=compte.id,
            montant=serializer.validated_data["montant"],
            user=request.user,
        )

        transaction.on_commit(
            lambda tx=tx: send_compte_depot_notification(tx)
        )

        compte.refresh_from_db()

        return Response(
            {
                "message": "Dépôt effectué avec succès.",

                "client": {
                    "nom": compte.client.nom,
                    "prenom": compte.client.prenom,
                    "telephone": compte.client.telephone,
                },

                "bijouterie": {
                    "id": client_bijouterie.id,
                    "nom": client_bijouterie.nom,
                },

                "transaction":
                CompteDepotTransactionSerializer(tx).data,

                "compte": {
                    "numero_compte": compte.numero_compte,
                    "nouveau_solde": str(compte.solde),
                },

                "receipt_url":
                request.build_absolute_uri(
                    f"/api/compte-depot/transactions/"
                    f"{tx.id}/receipt/80mm/"
                ),
            },
            status=status.HTTP_201_CREATED,
        )


# =========================================================
# RETRAIT
# =========================================================
class RetraitView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Effectuer un retrait sur un compte dépôt",
        operation_description="""
        Effectue un retrait sur un compte dépôt via téléphone client.

        Règles :
        - Admin : toutes les bijouteries
        - Manager : uniquement ses bijouteries
        - Caissier : uniquement sa bijouterie
        - Vendor : non autorisé
        - Buyer : non autorisé
        """,
        request_body=CompteDepotTelephoneTransactionSerializer,
        responses={
            201: openapi.Response(
                description="Retrait effectué avec succès"
            ),
            400: openapi.Response(
                description="Données invalides"
            ),
            403: openapi.Response(
                description="Accès refusé"
            ),
            404: openapi.Response(
                description="Compte introuvable"
            ),
        },
        tags=["compte dépôt"],
    )
    @transaction.atomic
    def post(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {
                    "message": (
                        "Seul l'admin, le manager ou le caissier "
                        "peut effectuer un retrait."
                    )
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = CompteDepotTelephoneTransactionSerializer(
            data=request.data
        )
        serializer.is_valid(raise_exception=True)

        telephone = serializer.validated_data["telephone"]

        try:
            compte = (
                CompteDepot.objects
                .select_for_update()
                .select_related(
                    "client",
                    "client__bijouterie",
                )
                .get(client__telephone=telephone)
            )
        except CompteDepot.DoesNotExist:
            return Response(
                {"detail": "Compte introuvable."},
                status=status.HTTP_404_NOT_FOUND
            )

        client_bijouterie = getattr(
            compte.client,
            "bijouterie",
            None,
        )

        if not client_bijouterie:
            return Response(
                {
                    "message":
                    "Ce compte dépôt n'est lié à aucune bijouterie."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if not user_can_access_bijouterie(
            request.user,
            client_bijouterie
        ):
            return Response(
                {
                    "message":
                    "Vous ne pouvez pas effectuer un retrait "
                    "sur ce compte."
                },
                status=status.HTTP_403_FORBIDDEN
            )

        tx = effectuer_retrait(
            compte_id=compte.id,
            montant=serializer.validated_data["montant"],
            user=request.user,
        )

        transaction.on_commit(
            lambda tx=tx: send_compte_depot_notification(tx)
        )

        compte.refresh_from_db()

        return Response(
            {
                "message": "Retrait effectué avec succès.",

                "client": {
                    "nom": compte.client.nom,
                    "prenom": compte.client.prenom,
                    "telephone": compte.client.telephone,
                },

                "bijouterie": {
                    "id": client_bijouterie.id,
                    "nom": client_bijouterie.nom,
                },

                "transaction":
                CompteDepotTransactionSerializer(tx).data,

                "compte": {
                    "numero_compte": compte.numero_compte,
                    "nouveau_solde": str(compte.solde),
                },

                "receipt_url":
                request.build_absolute_uri(
                    f"/api/compte-depot/transactions/"
                    f"{tx.id}/receipt/80mm/"
                ),
            },
            status=status.HTTP_201_CREATED
        )
        
# =========================================================
# SOLDE
# =========================================================
class GetSoldeAPIView(APIView):
    permission_classes = [IsAuthenticated]
    renderer_classes = [UserRenderer]

    @swagger_auto_schema(
        operation_summary="Consulter le solde d’un compte dépôt",
        operation_description="""
        Récupérer le solde d’un compte dépôt via le téléphone du client.

        Règles :
        - Admin : tous les comptes
        - Manager : uniquement ses bijouteries
        - Caissier : uniquement sa bijouterie
        - Vendor : uniquement sa bijouterie
        """,
        manual_parameters=[
            openapi.Parameter(
                "telephone",
                openapi.IN_QUERY,
                description="Numéro de téléphone du client lié au compte dépôt",
                type=openapi.TYPE_STRING,
                required=True,
            )
        ],
        responses={
            200: openapi.Response(description="Solde récupéré avec succès"),
            400: openapi.Response(description="Paramètre invalide ou compte sans bijouterie"),
            403: openapi.Response(description="Accès refusé"),
            404: openapi.Response(description="Compte non trouvé"),
        },
        tags=["compte dépôt"],
    )
    def get(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
            ROLE_VENDOR,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        telephone = request.query_params.get("telephone")

        if not telephone:
            return Response(
                {"detail": "Le paramètre 'telephone' est requis."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            compte = (
                CompteDepot.objects
                .select_related("client", "client__bijouterie")
                .get(client__telephone=telephone)
            )
        except CompteDepot.DoesNotExist:
            return Response(
                {"detail": "Compte non trouvé."},
                status=status.HTTP_404_NOT_FOUND,
            )

        client_bijouterie = getattr(
            compte.client,
            "bijouterie",
            None,
        )

        if not client_bijouterie:
            return Response(
                {"detail": "Ce compte n'est lié à aucune bijouterie."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not user_can_access_bijouterie(
            request.user,
            client_bijouterie,
        ):
            return Response(
                {
                    "detail":
                    "Vous n'avez pas accès à la bijouterie de ce compte."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        return Response(
            {
                "client": {
                    "nom": compte.client.nom,
                    "prenom": compte.client.prenom,
                    "telephone": compte.client.telephone,
                },
                "bijouterie": {
                    "id": client_bijouterie.id,
                    "nom": client_bijouterie.nom,
                },
                "compte": {
                    "numero_compte": compte.numero_compte,
                    "solde": str(compte.solde),
                    "created_at": compte.created_at,
                },
            },
            status=status.HTTP_200_OK,
        )
        
# =========================================================
# LISTE COMPTES AVEC FILTRE
# =========================================================
class ListerTousComptesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Lister les comptes dépôt",
        operation_description="""
        Liste les comptes dépôt accessibles à l'utilisateur.

        Règles :
        - Admin : tous les comptes
        - Manager : uniquement ses bijouteries
        - Caissier : uniquement sa bijouterie
        - Vendor : uniquement sa bijouterie
        """,
        manual_parameters=[
            openapi.Parameter(
                "telephone",
                openapi.IN_QUERY,
                description="Numéro de téléphone partiel ou complet",
                type=openapi.TYPE_STRING,
                required=False,
            )
        ],
        responses={
            200: openapi.Response(
                "Liste des comptes",
                CompteDepotSerializer(many=True),
            ),
            403: openapi.Response(
                description="Accès refusé"
            ),
        },
        tags=["compte dépôt"],
    )
    def get(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
            ROLE_VENDOR,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # ==========================================
        # BIJOUTERIES ACCESSIBLES
        # ==========================================

        bijouteries = get_user_bijouteries(request.user)

        if not bijouteries.exists():
            return Response(
                {
                    "detail":
                    "Aucune bijouterie accessible pour cet utilisateur."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # ==========================================
        # COMPTES ACCESSIBLES
        # ==========================================

        comptes = (
            CompteDepot.objects
            .select_related(
                "client",
                "client__bijouterie",
                "created_by",
            )
            .filter(
                client__bijouterie__in=bijouteries
            )
        )

        # ==========================================
        # FILTRE TELEPHONE
        # ==========================================

        telephone = request.query_params.get("telephone")

        if telephone:
            comptes = comptes.filter(
                client__telephone__icontains=telephone
            )

        comptes = comptes.order_by("-created_at")

        serializer = CompteDepotSerializer(
            comptes,
            many=True,
        )

        return Response(
            serializer.data,
            status=status.HTTP_200_OK,
        )
    
    
# =========================================================
# LISTE TRANSACTIONS
# =========================================================
class ListerToutesCompteDepotTransactionsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Lister les transactions compte dépôt",
        operation_description=(
            "Liste les transactions compte dépôt accessibles à l'utilisateur.\n\n"
            "### Règles\n"
            "- Admin : toutes les bijouteries.\n"
            "- Manager : uniquement ses bijouteries.\n"
            "- Caissier : uniquement sa bijouterie.\n"
        ),
        manual_parameters=[
            openapi.Parameter(
                "telephone",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "numero_compte",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "type_transaction",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "statut",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "start_date",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
            ),
            openapi.Parameter(
                "end_date",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
            ),
        ],
        responses={
            200: openapi.Response(
                "Liste des transactions",
                CompteDepotTransactionSerializer(many=True),
            ),
            403: openapi.Response(
                description="Accès refusé",
            ),
        },
        tags=["compte dépôt"],
    )
    def get(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        bijouteries = get_user_bijouteries(request.user)

        if not bijouteries.exists():
            return Response(
                {"detail": "Aucune bijouterie accessible."},
                status=status.HTTP_403_FORBIDDEN,
            )

        qs = (
            CompteDepotTransaction.objects
            .select_related(
                "compte",
                "compte__client",
                "compte__client__bijouterie",
                "user",
            )
            .filter(
                compte__client__bijouterie__in=bijouteries
            )
            .order_by("-date_transaction")
        )

        telephone = request.query_params.get("telephone")
        numero_compte = request.query_params.get("numero_compte")
        type_transaction = request.query_params.get("type_transaction")
        statut = request.query_params.get("statut")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if telephone:
            qs = qs.filter(
                compte__client__telephone__icontains=telephone
            )

        if numero_compte:
            qs = qs.filter(
                compte__numero_compte__icontains=numero_compte
            )

        if type_transaction:
            qs = qs.filter(
                type_transaction=type_transaction
            )

        if statut:
            qs = qs.filter(
                statut=statut
            )

        if start_date:
            qs = qs.filter(
                date_transaction__date__gte=start_date
            )

        if end_date:
            qs = qs.filter(
                date_transaction__date__lte=end_date
            )

        serializer = CompteDepotTransactionSerializer(
            qs,
            many=True,
        )

        return Response(
            serializer.data,
            status=status.HTTP_200_OK,
        )
        
# =========================================================
# EXPORT EXCEL TRANSACTIONS
# =========================================================

# historique mouvements
class ExportCompteDepotTransactionsExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Exporter les transactions compte dépôt en Excel",
        operation_description=(
            "Exporte les transactions compte dépôt accessibles à l'utilisateur.\n\n"
            "### Règles\n"
            "- Admin : toutes les bijouteries.\n"
            "- Manager : uniquement ses bijouteries.\n"
            "- Caissier : uniquement sa bijouterie.\n\n"
            "Filtres possibles : téléphone, numéro compte, type, statut, dates."
        ),
        manual_parameters=[
            openapi.Parameter(
                "telephone",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "numero_compte",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "type_transaction",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "statut",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                required=False,
            ),
            openapi.Parameter(
                "start_date",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
            ),
            openapi.Parameter(
                "end_date",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
            ),
        ],
        responses={
            200: openapi.Response(
                description="Fichier Excel généré avec succès"
            ),
            403: openapi.Response(
                description="Accès refusé"
            ),
        },
        tags=["compte dépôt"],
    )
    def get(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        bijouteries = get_user_bijouteries(request.user)

        if not bijouteries.exists():
            return Response(
                {"detail": "Aucune bijouterie accessible."},
                status=status.HTTP_403_FORBIDDEN,
            )

        qs = (
            CompteDepotTransaction.objects
            .select_related(
                "compte",
                "compte__client",
                "compte__client__bijouterie",
                "user",
            )
            .filter(
                compte__client__bijouterie__in=bijouteries
            )
            .order_by("-date_transaction")
        )

        telephone = request.query_params.get("telephone")
        numero_compte = request.query_params.get("numero_compte")
        type_transaction = request.query_params.get("type_transaction")
        statut = request.query_params.get("statut")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if telephone:
            qs = qs.filter(
                compte__client__telephone__icontains=telephone
            )

        if numero_compte:
            qs = qs.filter(
                compte__numero_compte__icontains=numero_compte
            )

        if type_transaction:
            qs = qs.filter(
                type_transaction=type_transaction
            )

        if statut:
            qs = qs.filter(
                statut=statut
            )

        if start_date:
            qs = qs.filter(
                date_transaction__date__gte=start_date
            )

        if end_date:
            qs = qs.filter(
                date_transaction__date__lte=end_date
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions Compte Depot"

        headers = [
            "Date",
            "Bijouterie",
            "Type",
            "Statut",
            "Numero compte",
            "Nom",
            "Prenom",
            "Telephone",
            "Montant",
            "Solde avant",
            "Solde apres",
            "Reference",
            "Commentaire",
            "Utilisateur",
        ]
        ws.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for tx in qs:
            client = getattr(tx.compte, "client", None)
            bijouterie = getattr(client, "bijouterie", None) if client else None

            user_label = ""

            if tx.user:
                user_label = (
                    getattr(tx.user, "email", None)
                    or getattr(tx.user, "username", "")
                    or str(tx.user)
                )

            ws.append([
                timezone.localtime(
                    tx.date_transaction
                ).strftime("%Y-%m-%d %H:%M:%S")
                if tx.date_transaction else "",

                getattr(bijouterie, "nom", "") if bijouterie else "",

                tx.get_type_transaction_display(),
                tx.get_statut_display(),
                tx.compte.numero_compte,

                getattr(client, "nom", "") if client else "",
                getattr(client, "prenom", "") if client else "",
                getattr(client, "telephone", "") if client else "",

                float(tx.montant),
                float(tx.solde_avant),
                float(tx.solde_apres),

                tx.reference or "",
                tx.commentaire or "",
                user_label,
            ])

        for column_cells in ws.columns:
            length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = min(length + 2, 40)

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        filename = (
            "transactions_compte_depot_"
            f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        wb.save(response)

        return response
    
    
# sauvegarde état des soldes actuels
class ExportCompteDepotSoldesExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Exporter les soldes des comptes dépôt en Excel",
        operation_description=(
            "Exporte l'état actuel des soldes des comptes dépôt.\n\n"
            "### Règles\n"
            "- Admin : toutes les bijouteries.\n"
            "- Manager : uniquement ses bijouteries.\n"
            "- Caissier : uniquement sa bijouterie.\n"
            "- Vendor : non autorisé.\n"
            "- Buyer : non autorisé."
        ),
        responses={
            200: openapi.Response(
                description="Fichier Excel généré avec succès"
            ),
            403: openapi.Response(
                description="Accès refusé"
            ),
        },
        tags=["compte dépôt"],
    )
    def get(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # ==========================================
        # BIJOUTERIES ACCESSIBLES
        # ==========================================

        bijouteries = get_user_bijouteries(request.user)

        if not bijouteries.exists():
            return Response(
                {
                    "detail":
                    "Aucune bijouterie accessible pour cet utilisateur."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # ==========================================
        # COMPTES ACCESSIBLES
        # ==========================================

        comptes = (
            CompteDepot.objects
            .select_related(
                "client",
                "client__bijouterie",
                "created_by",
            )
            .filter(
                client__bijouterie__in=bijouteries
            )
            .order_by(
                "client__bijouterie__nom",
                "client__nom",
            )
        )

        # ==========================================
        # EXCEL
        # ==========================================

        wb = Workbook()
        ws = wb.active
        ws.title = "Soldes Comptes Depot"

        headers = [
            "Date sauvegarde",
            "Bijouterie",
            "Nom client",
            "Prénom client",
            "Téléphone",
            "Numéro compte",
            "Solde actuel",
            "Date création compte",
            "Dernière transaction",
            "Créé par",
        ]

        ws.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        now = timezone.localtime(timezone.now())

        # ==========================================
        # DONNEES
        # ==========================================

        for compte in comptes:
            client = compte.client

            bijouterie = (
                getattr(client, "bijouterie", None)
                if client
                else None
            )

            # Dernière transaction du compte
            derniere_tx = (
                CompteDepotTransaction.objects
                .filter(compte=compte)
                .order_by("-date_transaction")
                .first()
            )

            created_by = ""

            if compte.created_by:
                created_by = (
                    getattr(compte.created_by, "email", None)
                    or getattr(compte.created_by, "username", "")
                    or str(compte.created_by)
                )

            ws.append([
                now.strftime("%Y-%m-%d %H:%M:%S"),

                getattr(bijouterie, "nom", "")
                if bijouterie
                else "",

                getattr(client, "nom", "")
                if client
                else "",

                getattr(client, "prenom", "")
                if client
                else "",

                getattr(client, "telephone", "")
                if client
                else "",

                compte.numero_compte,

                float(compte.solde),

                (
                    timezone.localtime(
                        compte.created_at
                    ).strftime("%Y-%m-%d %H:%M:%S")
                    if compte.created_at
                    else ""
                ),

                (
                    timezone.localtime(
                        derniere_tx.date_transaction
                    ).strftime("%Y-%m-%d %H:%M:%S")
                    if derniere_tx
                    else ""
                ),

                created_by,
            ])

        # ==========================================
        # LARGEUR COLONNES
        # ==========================================

        for column_cells in ws.columns:
            length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = min(length + 2, 45)

        # ==========================================
        # RESPONSE
        # ==========================================

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        filename = (
            "sauvegarde_soldes_compte_depot_"
            f"{now.strftime('%Y_%m_%d_%H%M%S')}.xlsx"
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )

        wb.save(response)

        return response
    
    
# =========================================================
# DASHBOARD COMPTE DEPOT
# =========================================================
class CompteDepotDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Dashboard comptes dépôt",
        operation_description=(
            "Dashboard des comptes dépôt accessibles à l'utilisateur.\n\n"
            "### Règles\n"
            "- Admin : toutes les bijouteries.\n"
            "- Manager : uniquement ses bijouteries.\n"
            "- Caissier : uniquement sa bijouterie.\n"
            "- Vendor : non autorisé.\n"
            "- Buyer : non autorisé.\n"
        ),
        manual_parameters=[
            openapi.Parameter(
                "start_date",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
            ),
            openapi.Parameter(
                "end_date",
                openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
            ),
        ],
        responses={
            200: openapi.Response(
                description="Dashboard comptes dépôt"
            ),
            403: openapi.Response(
                description="Accès refusé"
            ),
        },
        tags=["compte dépôt"],
    )
    def get(self, request):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        bijouteries = get_user_bijouteries(request.user)

        if not bijouteries.exists():
            return Response(
                {"detail": "Aucune bijouterie accessible."},
                status=status.HTTP_403_FORBIDDEN,
            )

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        compte_qs = (
            CompteDepot.objects
            .select_related(
                "client",
                "client__bijouterie",
            )
            .filter(
                client__bijouterie__in=bijouteries
            )
        )

        tx_qs = (
            CompteDepotTransaction.objects
            .select_related(
                "compte",
                "compte__client",
                "compte__client__bijouterie",
                "user",
            )
            .filter(
                compte__client__bijouterie__in=bijouteries
            )
        )

        if start_date:
            tx_qs = tx_qs.filter(
                date_transaction__date__gte=start_date
            )

        if end_date:
            tx_qs = tx_qs.filter(
                date_transaction__date__lte=end_date
            )

        total_comptes = compte_qs.count()

        total_solde_global = compte_qs.aggregate(
            total=Coalesce(
                Sum("solde"),
                Decimal("0.00"),
            )
        )["total"]

        total_depots = tx_qs.filter(
            type_transaction=CompteDepotTransaction.TYPE_DEPOT,
            statut=CompteDepotTransaction.STAT_TERMINE,
        ).aggregate(
            total=Coalesce(
                Sum("montant"),
                Decimal("0.00"),
            )
        )["total"]

        total_retraits = tx_qs.filter(
            type_transaction=CompteDepotTransaction.TYPE_RETRAIT,
            statut=CompteDepotTransaction.STAT_TERMINE,
        ).aggregate(
            total=Coalesce(
                Sum("montant"),
                Decimal("0.00"),
            )
        )["total"]

        nombre_transactions = tx_qs.count()

        top_comptes = (
            compte_qs
            .order_by("-solde")[:10]
        )

        top_comptes_data = []

        for compte in top_comptes:
            client = getattr(compte, "client", None)
            bijouterie = (
                getattr(client, "bijouterie", None)
                if client
                else None
            )

            top_comptes_data.append({
                "numero_compte": compte.numero_compte,
                "solde": compte.solde,
                "client_nom": getattr(client, "nom", "") if client else "",
                "client_prenom": getattr(client, "prenom", "") if client else "",
                "telephone": getattr(client, "telephone", "") if client else "",
                "bijouterie": {
                    "id": bijouterie.id,
                    "nom": bijouterie.nom,
                } if bijouterie else None,
            })

        transactions_par_type = (
            tx_qs
            .values("type_transaction")
            .annotate(
                count=Count("id"),
                total=Coalesce(
                    Sum("montant"),
                    Decimal("0.00"),
                ),
            )
            .order_by("type_transaction")
        )

        latest_transactions = (
            tx_qs
            .order_by("-date_transaction")[:10]
        )

        return Response(
            {
                "periode": {
                    "start_date": start_date,
                    "end_date": end_date,
                },

                "kpis": {
                    "total_comptes": total_comptes,
                    "total_solde_global": total_solde_global,
                    "total_depots": total_depots,
                    "total_retraits": total_retraits,
                    "nombre_transactions": nombre_transactions,
                },

                "transactions_par_type": list(
                    transactions_par_type
                ),

                "top_comptes": top_comptes_data,

                "dernieres_transactions":
                CompteDepotTransactionSerializer(
                    latest_transactions,
                    many=True,
                ).data,
            },
            status=status.HTTP_200_OK,
        )
        
# =========================================================
# RECU PDF 80MM
# =========================================================
class CompteDepotTransactionReceipt80mmPDFAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_id="compteDepotTransactionReceipt",
        operation_summary="Récupérer les données d'un reçu compte dépôt",
        operation_description=(
            "Retourne les données nécessaires au frontend pour afficher "
            "et imprimer un reçu compte dépôt.\n\n"
            "### Règles\n"
            "- Admin : toutes les bijouteries.\n"
            "- Manager : uniquement ses bijouteries.\n"
            "- Caissier : uniquement sa bijouterie.\n"
            "- Vendor : non autorisé.\n"
            "- Buyer : non autorisé."
        ),
        responses={
            200: openapi.Response(
                description="Données du reçu récupérées avec succès"
            ),
            400: openapi.Response(
                description="Transaction sans bijouterie"
            ),
            403: openapi.Response(
                description="Accès refusé"
            ),
            404: openapi.Response(
                description="Transaction introuvable"
            ),
        },
        tags=["compte dépôt"],
    )
    def get(self, request, transaction_id):
        role = get_role_name(request.user)

        if role not in {
            ROLE_ADMIN,
            ROLE_MANAGER,
            ROLE_CASHIER,
        }:
            return Response(
                {"message": "Accès refusé."},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            tx = (
                CompteDepotTransaction.objects
                .select_related(
                    "compte",
                    "compte__client",
                    "compte__client__bijouterie",
                    "user",
                )
                .get(pk=transaction_id)
            )

        except CompteDepotTransaction.DoesNotExist:
            return Response(
                {"detail": "Transaction introuvable."},
                status=status.HTTP_404_NOT_FOUND,
            )

        compte = tx.compte
        client = compte.client

        client_bijouterie = getattr(
            client,
            "bijouterie",
            None,
        )

        if not client_bijouterie:
            return Response(
                {
                    "detail":
                    "Cette transaction n'est liée à aucune bijouterie."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not user_can_access_bijouterie(
            request.user,
            client_bijouterie,
        ):
            return Response(
                {
                    "detail":
                    "Vous n'avez pas accès à cette transaction."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        return Response(
            {
                "organisation": {
                    "nom": "BIJOUTERIE RIO GOLD",
                    "bijouterie": {
                        "id": client_bijouterie.id,
                        "nom": client_bijouterie.nom,
                    },
                },

                "client": {
                    "nom": client.nom,
                    "prenom": client.prenom,
                    "telephone": client.telephone,
                },

                "compte": {
                    "id": compte.id,
                    "numero_compte": compte.numero_compte,
                    "solde": str(compte.solde),
                },

                "transaction": {
                    "id": tx.id,
                    "type_transaction": tx.type_transaction,
                    "type_transaction_label":
                    tx.get_type_transaction_display(),

                    "montant": str(tx.montant),

                    "solde_avant": str(tx.solde_avant),
                    "solde_apres": str(tx.solde_apres),

                    "statut": tx.statut,
                    "statut_label":
                    tx.get_statut_display(),

                    "reference": tx.reference,
                    "commentaire": tx.commentaire,

                    "date_transaction":
                    tx.date_transaction,
                },

                "effectue_par": {
                    "id": tx.user_id,
                    "email": (
                        getattr(tx.user, "email", None)
                        if tx.user
                        else None
                    ),
                    "username": (
                        getattr(tx.user, "username", None)
                        if tx.user
                        else None
                    ),
                },
            },
            status=status.HTTP_200_OK,
        )
        
        